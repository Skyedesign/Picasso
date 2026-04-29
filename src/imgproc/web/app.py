"""FastAPI app driving the local imgproc UI.

Endpoints are intentionally minimal — the UI is a thin shell around the same CLI
and config machinery that runs from the terminal. Processing happens in a
background thread and is polled via `/api/jobs/{id}`.

Path safety: every `batch_name` received from the client is validated to ensure
the resolved path stays inside `BATCHES_ROOT`. No path traversal.
"""

from __future__ import annotations

import asyncio
import base64
import os
import re
import shutil
import socket
import subprocess
import sys
import threading
import uuid
import webbrowser
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Literal

import yaml
from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from PIL import Image

from ..batch_meta import (
    BatchMeta,
    ImageRow,
    ImageVerdict,
    now_iso,
    read_meta,
    update_verdict,
    update_verdicts_bulk,
    write_meta,
)
from ..cli import IMAGE_EXTS, process_folder
from ..config import Config, find_project_root, load_config
from ..engine import detect_product, normalize_to_canvas
from ..engine.background import detect_background
from ..ingest.sortlib import (
    Anchor,
    Candidate,
    DupeCluster,
    Match,
    find_dupe_clusters,
    hash_candidates,
    load_anchors_from_xlsx,
    rank_candidates_per_sku,
)
from ..output import status_subfolder
from ..sheetcheck import (
    Suppressions,
    apply_suppressions,
    load_suffixes,
    parse_sheet,
    read_suppressions,
    run_rules,
    suppression_path,
    write_suppressions,
)
from ..updater import check_for_update, perform_swap

STATIC = Path(__file__).parent / "static"
_PROJECT_ROOT = find_project_root()
BATCHES_ROOT = (_PROJECT_ROOT / "batches").resolve()
CONFIG_PATH = _PROJECT_ROOT / "imgproc.yaml"
SUFFIXES_PATH = _PROJECT_ROOT / "picasso-suffixes.yaml"
# Scratch dir for editable working copies of xlsx files outside the
# batch flow — Sheet check's "Make editable copy" lands here so Apply
# fixes can be written without touching source/. Lives alongside
# batches/ at project root so it's visible in Explorer and easy to
# clean up; gitignored at the repo level (no checked-in working files).
SCRATCH_ROOT = (_PROJECT_ROOT / "scratch" / "sheetcheck").resolve()
BATCHES_ROOT.mkdir(exist_ok=True)

# Conservative: alphanumerics, dash, underscore, space. Prevents path traversal and
# OS-reserved characters on Windows.
_NAME_RE = re.compile(r"^[A-Za-z0-9 _\-]+$")

app = FastAPI(title="imgproc")
app.mount("/static", StaticFiles(directory=STATIC), name="static")
# Expose batches dir so the browser can load report.html and its assets directly,
# and so we can show original images as thumbnails in the UI.
app.mount("/batches", StaticFiles(directory=BATCHES_ROOT, html=True), name="batches")


def _resolve_batch(name: str) -> Path:
    if not _NAME_RE.match(name):
        raise HTTPException(400, "invalid batch name")
    folder = (BATCHES_ROOT / name).resolve()
    try:
        folder.relative_to(BATCHES_ROOT)
    except ValueError:
        raise HTTPException(400, "path escapes batches root")
    if not folder.is_dir():
        raise HTTPException(404, "batch not found")
    return folder


# ─── Pages ────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def index() -> str:
    return (STATIC / "index.html").read_text(encoding="utf-8")


@app.get("/demo", response_class=HTMLResponse)
def demo_page() -> str:
    """Standalone Demo Resizer at its own URL — for "open in new tab" so Alida
    can run the tool alongside the main UI without modal context-switching."""
    return (STATIC / "demo.html").read_text(encoding="utf-8")


@app.get("/reviewer/{name}", response_class=HTMLResponse)
def reviewer_page(name: str) -> str:
    """Visual reviewer for a single batch. Validates the name to match the
    batch-folder rules; the actual data is fetched client-side from
    /api/batches/{name}/state. Serving the same HTML for any valid batch
    keeps caching simple."""
    if not _NAME_RE.match(name):
        raise HTTPException(400, "invalid batch name")
    return (STATIC / "reviewer.html").read_text(encoding="utf-8")


@app.get("/sheetcheck", response_class=HTMLResponse)
def sheetcheck_page() -> str:
    """Sheet check (xlsx linter) page. The file picker + run button are
    JS-driven against the /api/sheetcheck/* endpoints below."""
    return (STATIC / "sheetcheck.html").read_text(encoding="utf-8")


@app.get("/sort/{name}", response_class=HTMLResponse)
def sort_page(name: str) -> str:
    """Visual SKU-matching tool for one batch. Loads anchors from a
    user-chosen xlsx + bbox-pHashes the batch's images + lets the user
    pick a hero per SKU. Output goes to {batch}/processed/sorted/.
    Same JS for any valid batch name; client fetches state via
    /api/sort/* endpoints."""
    if not _NAME_RE.match(name):
        raise HTTPException(400, "invalid batch name")
    return (STATIC / "sort.html").read_text(encoding="utf-8")


# ─── Batches ──────────────────────────────────────────────────────────────

@app.get("/api/batches")
def list_batches() -> dict:
    items = []
    for folder in sorted(BATCHES_ROOT.iterdir()):
        if not folder.is_dir():
            continue
        n_images = sum(
            1 for p in folder.iterdir()
            if p.is_file() and p.suffix.lower() in IMAGE_EXTS
        )
        has_report = (folder / "report.html").exists()
        processed_count = 0
        review_count = 0
        sorted_count = 0
        if (folder / "processed").is_dir():
            processed_count = sum(
                1 for p in (folder / "processed").iterdir()
                if p.is_file() and p.suffix.lower() in IMAGE_EXTS
            )
        sorted_dir = folder / "processed" / "sorted"
        if sorted_dir.is_dir():
            sorted_count = sum(
                1 for p in sorted_dir.iterdir()
                if p.is_file() and p.suffix.lower() in IMAGE_EXTS
            )
        if (folder / "review").is_dir():
            review_count = sum(
                1 for p in (folder / "review").iterdir()
                if p.suffix.lower() in IMAGE_EXTS
            )
        # Send-state fields come from the sidecar; absent for legacy /
        # never-sent batches (None ⇒ UI shows "Never sent").
        meta = read_meta(folder)
        xlsx_filename: str | None = None
        if meta and meta.xlsx_filename and (folder / meta.xlsx_filename).is_file():
            xlsx_filename = meta.xlsx_filename
        items.append({
            "name": folder.name,
            "image_count": n_images,
            "has_report": has_report,
            "processed_count": processed_count,
            "review_count": review_count,
            "sorted_count": sorted_count,
            "xlsx_filename": xlsx_filename,
            "last_sent_at": meta.last_sent_at if meta else None,
            "last_sent_count": meta.last_sent_count if meta else None,
            "pegasus_received_at": meta.pegasus_received_at if meta else None,
        })
    return {"root": str(BATCHES_ROOT), "batches": items}


class NewBatch(BaseModel):
    name: str


@app.post("/api/batches")
def create_batch(body: NewBatch) -> dict:
    name = body.name.strip()
    if not _NAME_RE.match(name):
        raise HTTPException(400, "name may only contain letters, numbers, dash, underscore, space")
    target = BATCHES_ROOT / name
    if target.exists():
        raise HTTPException(409, "batch already exists")
    target.mkdir()
    return {"name": name}


_SOURCE_MAX_DEPTH = 3  # how deep under source/ to walk when populating the dropdown


@app.get("/api/source-folders")
def list_source_folders() -> dict:
    """Enumerate folders under source/ that contain image files, so the UI can
    show Alida a picker instead of requiring her to paste a path."""
    source_root = (_PROJECT_ROOT / "source").resolve()
    if not source_root.is_dir():
        return {"root": str(source_root), "exists": False, "folders": []}

    folders = []

    def walk(path: Path, depth: int) -> None:
        if depth > _SOURCE_MAX_DEPTH:
            return
        try:
            entries = list(path.iterdir())
        except (PermissionError, OSError):
            return
        image_count = sum(
            1 for p in entries
            if p.is_file() and p.suffix.lower() in IMAGE_EXTS
        )
        if image_count > 0:
            folders.append({
                "path": str(path),
                "relative": str(path.relative_to(source_root)) or ".",
                "image_count": image_count,
            })
        for p in entries:
            if p.is_dir() and not p.name.startswith("."):
                walk(p, depth + 1)

    walk(source_root, 0)
    folders.sort(key=lambda f: f["relative"].lower())
    return {"root": str(source_root), "exists": True, "folders": folders}


class ImportRequest(BaseModel):
    name: str
    source_path: str
    move: bool = False  # if True, move the source files; otherwise copy
    # Optional xlsx to attach to the batch in one step. Always copied
    # (never moved) — keeping the source xlsx pristine is the whole
    # point of the cleanup-workspace pattern.
    xlsx_path: str | None = None


@app.post("/api/batches/import")
def import_folder(body: ImportRequest) -> dict:
    # Target: must be a fresh batch under BATCHES_ROOT.
    name = body.name.strip()
    if not _NAME_RE.match(name):
        raise HTTPException(400, "name may only contain letters, numbers, dash, underscore, space")
    target = BATCHES_ROOT / name
    if target.exists():
        raise HTTPException(409, "batch already exists — pick a different name")

    # Source: any folder on the filesystem (read-only from our side). Strip the
    # surrounding quotes Windows' "Copy as path" adds, and expand ~ for convenience.
    src_str = body.source_path.strip().strip('"').strip("'")
    if not src_str:
        raise HTTPException(400, "source folder path is required")
    source = Path(src_str).expanduser()
    if not source.exists():
        raise HTTPException(400, f"source folder not found: {source}")
    if not source.is_dir():
        raise HTTPException(400, f"source path is not a folder: {source}")

    # Prevent importing from inside BATCHES_ROOT itself — the only legitimate case
    # is cross-folder reorganization, which should use the existing batch-rename flow
    # (which doesn't exist yet; ask the user before copying inside the tree).
    try:
        source.resolve().relative_to(BATCHES_ROOT.resolve())
        raise HTTPException(400, "source is inside the batches folder; import from an outside location")
    except ValueError:
        pass  # good — source is outside BATCHES_ROOT

    # Threshold matches the lifestyle filter default — anything below this will be
    # routed to skipped/ at process time, so flagging it now lets Alida pre-cull.
    non_white_threshold = 0.85

    target.mkdir()
    imported = 0
    non_white: list[str] = []
    try:
        for p in source.iterdir():
            if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
                dest = target / p.name
                if body.move:
                    shutil.move(str(p), str(dest))
                else:
                    shutil.copy2(p, dest)
                imported += 1
                # Check the copied file's background. Cheap — samples 4 corners only.
                try:
                    with Image.open(dest) as img:
                        bg = detect_background(img)
                        if bg.purity < non_white_threshold:
                            non_white.append(p.name)
                except Exception:
                    pass  # skip malformed images silently — they'll surface at process time
    except Exception:
        # Clean up a half-imported batch so the user isn't left with a weird state.
        shutil.rmtree(target, ignore_errors=True)
        raise

    if imported == 0:
        target.rmdir()
        raise HTTPException(400, "no image files (.jpg, .png, .webp, .bmp) found in source folder")

    # Optional xlsx attach. Done last so an xlsx-validation failure doesn't
    # require unwinding the image import. Failure here is non-fatal: the
    # batch still has its images; the user can re-attach via the dedicated
    # endpoint later.
    xlsx_attached: str | None = None
    xlsx_error: str | None = None
    if body.xlsx_path:
        try:
            src_xlsx = _resolve_xlsx_path(body.xlsx_path)
            shutil.copy2(src_xlsx, target / src_xlsx.name)
            meta = read_meta(target) or BatchMeta(batch_name=name)
            meta.xlsx_filename = src_xlsx.name
            write_meta(target, meta)
            xlsx_attached = src_xlsx.name
        except HTTPException as e:
            xlsx_error = e.detail
        except OSError as e:
            xlsx_error = f"copy failed: {e}"

    return {
        "name": name,
        "imported": imported,
        "moved": body.move,
        "non_white_count": len(non_white),
        "non_white_files": non_white,
        "xlsx_attached": xlsx_attached,
        "xlsx_error": xlsx_error,
    }


@app.delete("/api/batches/{name}")
def delete_batch(name: str) -> dict:
    # Destructive — recursively removes the batch folder and everything it contains
    # (originals, processed/, review/, report, assets). The `_resolve_batch` helper
    # ensures `name` can't escape BATCHES_ROOT via path traversal.
    folder = _resolve_batch(name)
    shutil.rmtree(folder)
    return {"ok": True, "deleted": name}


def _batch_xlsx_path(folder: Path, meta: BatchMeta | None = None) -> Path | None:
    """Where this batch's working xlsx lives, if any.

    Sidecar-tracked: `BatchMeta.xlsx_filename` is the canonical record.
    Returns None for batches without an attached xlsx (legacy + un-set).
    Re-reads the sidecar when caller doesn't provide one.
    """
    if meta is None:
        meta = read_meta(folder)
    if not meta or not meta.xlsx_filename:
        return None
    p = folder / meta.xlsx_filename
    if not p.is_file():
        return None
    return p


class AttachXlsxRequest(BaseModel):
    xlsx_path: str
    overwrite: bool = False


@app.post("/api/batches/{name}/attach-xlsx")
def attach_xlsx(name: str, body: AttachXlsxRequest) -> dict:
    """Copy an external xlsx into this batch as the working copy.

    Stores under `{batch}/{xlsx_filename}` — name preserved so Alida
    recognises it on disk. The sidecar's `xlsx_filename` field becomes
    the source of truth for "this batch has an xlsx"; downstream
    features (Sheet check write-back, Visual Sort, Send) read it.

    Refuses to overwrite an existing attached xlsx unless `overwrite`
    is true — the on-disk copy may have edits Alida hasn't synced
    elsewhere yet.
    """
    folder = _resolve_batch(name)
    src = _resolve_xlsx_path(body.xlsx_path)

    meta = read_meta(folder) or BatchMeta(batch_name=name)
    if meta.xlsx_filename and not body.overwrite:
        existing = folder / meta.xlsx_filename
        if existing.is_file():
            raise HTTPException(
                409,
                f"batch already has {meta.xlsx_filename}; pass overwrite=true to replace",
            )

    dest = folder / src.name
    try:
        shutil.copy2(src, dest)
    except OSError as e:
        raise HTTPException(500, f"copy failed: {e}")

    meta.xlsx_filename = src.name
    write_meta(folder, meta)
    return {"ok": True, "xlsx_filename": src.name, "path": str(dest)}


_REPEAT_SUFFIX_RE = re.compile(r"^(?P<base>.+?)-v(?P<n>\d+)$")


def _next_repeat_name(base_name: str) -> str:
    """Pick the first free `{root}-vN` slot.

    If `base_name` is `flowers`, the first sibling is `flowers-v2`,
    next `flowers-v3`, …. If it already ends in `-vN` we keep the
    same root and increment, so repeating `flowers-v2` produces
    `flowers-v3` — not `flowers-v2-v2`. Up to v999 just so the loop
    terminates; nobody's going to repeat 1000 times.
    """
    m = _REPEAT_SUFFIX_RE.match(base_name)
    if m:
        root = m.group("base")
        start = int(m.group("n")) + 1
    else:
        root = base_name
        start = 2
    for n in range(start, 1000):
        cand = f"{root}-v{n}"
        if not (BATCHES_ROOT / cand).exists():
            return cand
    raise HTTPException(500, "ran out of repeat suffixes (>=1000); time to clean up")


@app.post("/api/batches/{name}/repeat")
def repeat_batch(name: str) -> dict:
    """Create a sibling batch with the source's originals + config.

    Copies (never moves):
      • top-level image originals (the inputs, not processing output)
      • `folder.yaml` if present (preserves per-batch tuning)
      • the attached xlsx if any (so the new run starts from the same
        cleaned spreadsheet — re-doing the visual sort is the
        expected use case)

    Skips: `processed/`, `review/`, `skipped/`, `report.html`,
    `_report_assets/`, `batch.json`. Those are outputs of the previous
    run; the new batch starts fresh and will re-create them on Process.

    The source batch is NOT modified.
    """
    src_folder = _resolve_batch(name)
    new_name = _next_repeat_name(name)
    dst_folder = BATCHES_ROOT / new_name
    if dst_folder.exists():  # paranoia — _next_repeat_name should have prevented this
        raise HTTPException(409, f"sibling {new_name} already exists")

    dst_folder.mkdir()
    n_images = 0
    n_xlsx = 0
    try:
        # Top-level originals only.
        for p in src_folder.iterdir():
            if not p.is_file():
                continue
            if p.suffix.lower() in IMAGE_EXTS:
                shutil.copy2(p, dst_folder / p.name)
                n_images += 1
            elif p.name == "folder.yaml":
                shutil.copy2(p, dst_folder / p.name)
        # Attached xlsx — read filename from sidecar (single source of truth)
        # rather than guessing by extension, since other xlsx files near
        # the batch wouldn't be ours to copy.
        src_meta = read_meta(src_folder)
        xlsx_filename: str | None = None
        if src_meta and src_meta.xlsx_filename:
            xlsx_src = src_folder / src_meta.xlsx_filename
            if xlsx_src.is_file():
                shutil.copy2(xlsx_src, dst_folder / xlsx_src.name)
                xlsx_filename = xlsx_src.name
                n_xlsx = 1
    except Exception:
        # Roll back the half-built sibling so the user isn't left with a
        # phantom empty batch.
        shutil.rmtree(dst_folder, ignore_errors=True)
        raise

    if n_images == 0:
        # Empty source ⇒ pointless to repeat. Clean up + 400 so the UI
        # can show a clear message.
        shutil.rmtree(dst_folder, ignore_errors=True)
        raise HTTPException(
            400,
            f"batch {name} has no top-level images to repeat",
        )

    # Sidecar for the new batch: only carry forward xlsx_filename + the
    # last_run_config snapshot (so the user can see what knobs the
    # previous run used). Leave verdicts, run timestamp, send-state out
    # — the new batch hasn't done any of those things yet.
    new_meta = BatchMeta(batch_name=new_name)
    if xlsx_filename:
        new_meta.xlsx_filename = xlsx_filename
    if src_meta and src_meta.last_run_config:
        new_meta.last_run_config = dict(src_meta.last_run_config)
    write_meta(dst_folder, new_meta)

    return {
        "ok": True,
        "name": new_name,
        "image_count": n_images,
        "xlsx_attached": bool(n_xlsx),
        "from": name,
    }


@app.post("/api/batches/{name}/open")
def open_in_explorer(name: str) -> dict:
    folder = _resolve_batch(name)
    # Windows-only for now; the user's environment is Win11.
    if sys.platform == "win32":
        subprocess.Popen(["explorer", str(folder)])
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(folder)])
    else:
        subprocess.Popen(["xdg-open", str(folder)])
    return {"ok": True}


# ─── Config ───────────────────────────────────────────────────────────────

@app.get("/api/config")
def get_config() -> dict:
    if CONFIG_PATH.exists():
        data = yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8")) or {}
    else:
        data = Config().model_dump()
    # Round-trip through Config to return validated + defaulted values.
    cfg = Config(**data)
    out = cfg.model_dump()
    out["output_canvas"] = list(out["output_canvas"])
    return out


@app.post("/api/config")
def save_config(body: dict[str, Any]) -> dict:
    # Validate via the same pydantic schema the CLI uses.
    cfg = Config(**body)
    data = cfg.model_dump()
    data["output_canvas"] = list(data["output_canvas"])
    CONFIG_PATH.write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    return {"ok": True}


# ─── Processing jobs ──────────────────────────────────────────────────────

_jobs: dict[str, dict] = {}
_jobs_lock = threading.Lock()


class ProcessRequest(BaseModel):
    name: str
    tolerance: float | None = None
    target_ratio: float | None = None
    dry_run: bool = False


@app.post("/api/process")
def start_processing(body: ProcessRequest) -> dict:
    folder = _resolve_batch(body.name)

    job_id = uuid.uuid4().hex
    with _jobs_lock:
        _jobs[job_id] = {
            "status": "running",
            "log": "",
            "batch": body.name,
            "report_url": None,
            "progress": {"phase": "starting", "current": 0, "total": 0},
        }

    overrides: dict = {}
    if body.tolerance is not None:
        overrides["tolerance_mad"] = body.tolerance
    if body.target_ratio is not None:
        overrides["target_ratio"] = body.target_ratio

    def _log(msg: str) -> None:
        with _jobs_lock:
            cur = _jobs[job_id]["log"]
            _jobs[job_id]["log"] = (cur + "\n" + msg) if cur else msg

    def _progress(p: dict) -> None:
        # Snapshot under the lock — UI polls and reads atomically.
        with _jobs_lock:
            _jobs[job_id]["progress"] = dict(p)

    def run() -> None:
        try:
            result = process_folder(
                folder,
                cli_overrides=overrides,
                progress=_progress,
                log=_log,
                dry_run=body.dry_run,
            )
            ok = bool(result.get("ok"))
        except Exception as e:
            _log(f"Exception: {e}")
            ok = False

        with _jobs_lock:
            _jobs[job_id]["status"] = "done" if ok else "error"
            if ok and (folder / "report.html").exists():
                _jobs[job_id]["report_url"] = f"/batches/{body.name}/report.html"

    threading.Thread(target=run, daemon=True).start()
    return {"job_id": job_id}


@app.get("/api/jobs/{job_id}")
def get_job(job_id: str) -> dict:
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            raise HTTPException(404, "job not found")
        return dict(job)


# ─── Demo Resizer & batch image picker ───────────────────────────────────
# Demo Resizer is the "tune the fill ratio on one image" tool, ported from
# the Pegasus-embedded modal. Standalone Picasso calls the engine in-process
# (no subprocess to imgproc-hero), wraps in to_thread to keep handlers
# responsive while Pillow holds the GIL on a big TIFF decode.

def _render_with_overrides(
    img: Image.Image,
    cfg: Config,
    *,
    target_ratio: float,
    canvas_size: tuple[int, int],
    bbox: tuple[int, int, int, int] | None = None,
    centroid: tuple[float, float] | None = None,
    source_path: Path | None = None,
    padding_pct: float = 0.0,
) -> tuple[bytes, dict]:
    """Detect + apply optional bbox/centroid overrides + normalize.

    Returns (JPEG bytes, summary dict). Shared by Demo Resizer preview,
    reviewer live-preview, and the verdict re-run path so override
    semantics stay identical across all three.

    `padding_pct` defaults to 0 because every caller of this helper is in
    a per-image override workflow where target_ratio is the user's
    explicit request — letting the configured padding constrain the result
    would create a dead zone at the top of the slider. The CLI's group-
    batch path is unaffected; it calls `normalize_to_canvas` directly with
    `cfg.padding_pct`.
    """
    det = detect_product(source_path or Path(":preview:"), img, bg_threshold=cfg.bg_threshold)
    if det.bbox[2] - det.bbox[0] <= 0 or det.bbox[3] - det.bbox[1] <= 0:
        raise HTTPException(422, "no product detected")

    # bbox override: must be in-bounds with positive area.
    if bbox is not None:
        l, t, r, b = bbox
        if not (0 <= l < r <= det.width and 0 <= t < b <= det.height):
            raise HTTPException(422, "bbox out of bounds")
        det.bbox = (l, t, r, b)
        det.occupied_ratio = ((r - l) * (b - t)) / (det.width * det.height)
        if centroid is None:
            det.centroid = ((l + r) / 2.0, (t + b) / 2.0)

    # centroid override: a single click in the original sets where the
    # product lands on the canvas. Forces the mask-centroid path so
    # `det.centroid` is honoured regardless of the global recenter pref.
    recenter_on_mask = cfg.recenter
    if centroid is not None:
        cx, cy = centroid
        if not (0 <= cx <= det.width and 0 <= cy <= det.height):
            raise HTTPException(422, "centroid out of bounds")
        det.centroid = (float(cx), float(cy))
        recenter_on_mask = True
    if bbox is not None and centroid is None:
        recenter_on_mask = True

    normalized = normalize_to_canvas(
        det,
        target_ratio=target_ratio,
        canvas_size=canvas_size,
        padding_pct=padding_pct,
        max_upscale=cfg.max_upscale,
        recenter_on_mask_centroid=recenter_on_mask,
    )
    out = BytesIO()
    normalized.save(out, format="JPEG", quality=92)
    return out.getvalue(), {
        "occupied_ratio": det.occupied_ratio,
        "confidence": det.confidence,
        "bg_purity": det.bg.purity,
        "bg_is_white": det.bg.is_white,
    }


class PreviewRequest(BaseModel):
    image_b64: str
    ext: str = "jpg"
    target_ratio: float = Field(ge=0.0, le=1.0)
    canvas_size: tuple[int, int] | None = None  # falls back to imgproc.yaml's value
    bbox: tuple[int, int, int, int] | None = None
    centroid: tuple[float, float] | None = None


@app.post("/api/picasso/preview")
async def picasso_preview(body: PreviewRequest) -> dict:
    """Render a single image with overrides and return a JPEG as base64."""
    try:
        src_bytes = base64.b64decode(body.image_b64)
    except Exception:
        raise HTTPException(400, "bad base64")
    if not src_bytes:
        raise HTTPException(400, "empty image")

    cfg = load_config(CONFIG_PATH)
    canvas_size = tuple(body.canvas_size) if body.canvas_size else cfg.output_canvas
    ext = body.ext.lower().lstrip(".") or "jpg"

    def _render() -> bytes:
        try:
            img = Image.open(BytesIO(src_bytes))
            img.load()
        except Exception as e:
            raise HTTPException(400, f"cannot read image: {e}")
        out_bytes, _ = _render_with_overrides(
            img, cfg,
            target_ratio=body.target_ratio,
            canvas_size=canvas_size,
            bbox=body.bbox,
            centroid=body.centroid,
            source_path=Path(f"preview.{ext}"),
        )
        return out_bytes

    out_bytes = await asyncio.to_thread(_render)
    return {
        "image_b64": base64.b64encode(out_bytes).decode("ascii"),
        "ratio": body.target_ratio,
        "canvas_size": list(canvas_size),
    }


class BatchPreviewRequest(BaseModel):
    target_ratio: float = Field(ge=0.0, le=1.0)
    canvas_size: tuple[int, int] | None = None
    bbox: tuple[int, int, int, int] | None = None
    centroid: tuple[float, float] | None = None


@app.post("/api/batches/{name}/preview/{filename}")
async def picasso_preview_from_batch(name: str, filename: str, body: BatchPreviewRequest) -> dict:
    """Live preview for the reviewer's re-run panel: load the original from
    the batch's top-level, apply overrides, return the rendered JPEG as
    base64. Avoids round-tripping the image through base64 over the wire
    when it's already on disk in the batch."""
    folder = _resolve_batch(name)
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(400, "invalid filename")
    src_path = folder / filename
    if not src_path.is_file():
        raise HTTPException(404, f"source image not at top level: {filename}")

    cfg = load_config(CONFIG_PATH)
    canvas_size = tuple(body.canvas_size) if body.canvas_size else cfg.output_canvas

    def _render() -> bytes:
        try:
            img = Image.open(src_path)
            img.load()
        except Exception as e:
            raise HTTPException(400, f"cannot read image: {e}")
        out_bytes, _ = _render_with_overrides(
            img, cfg,
            target_ratio=body.target_ratio,
            canvas_size=canvas_size,
            bbox=body.bbox,
            centroid=body.centroid,
            source_path=src_path,
        )
        return out_bytes

    out_bytes = await asyncio.to_thread(_render)
    return {
        "image_b64": base64.b64encode(out_bytes).decode("ascii"),
        "ratio": body.target_ratio,
        "canvas_size": list(canvas_size),
    }


_SUBFOLDERS = {"processed", "review", "skipped"}


@app.get("/api/batches/{name}/images")
def list_batch_images(name: str) -> dict:
    """List image files in a batch grouped by location, for the Demo Resizer
    "pick from batch" picker. Returns top-level (originals), processed/, and
    review/ filenames only — no full paths (the picker calls the thumbnail
    endpoint by basename + sub)."""
    folder = _resolve_batch(name)
    out: dict = {"top": []}
    for p in sorted(folder.iterdir(), key=lambda x: x.name.lower()):
        if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
            out["top"].append(p.name)
    for sub in ("processed", "review"):
        sub_dir = folder / sub
        out[sub] = []
        if sub_dir.is_dir():
            for p in sorted(sub_dir.iterdir(), key=lambda x: x.name.lower()):
                if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
                    out[sub].append(p.name)
    return out


@app.get("/api/batches/{name}/thumbs/{filename}")
def batch_thumbnail(name: str, filename: str, w: int = 200, sub: str = "") -> Response:
    """Small JPEG thumbnail for the picker grid. Pillow's `draft` mode
    decodes JPEGs at a downsampled resolution — substantially faster than a
    full decode when the source is multi-MP."""
    folder = _resolve_batch(name)
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(400, "invalid filename")
    if sub and sub not in _SUBFOLDERS:
        raise HTTPException(400, "invalid subfolder")
    if w <= 0 or w > 1024:
        raise HTTPException(400, "invalid width")

    file_path = (folder / sub / filename) if sub else (folder / filename)
    if not file_path.is_file():
        raise HTTPException(404, "image not found")

    img = Image.open(file_path)
    if img.format == "JPEG":
        img.draft("RGB", (w * 2, w * 2))
    img.thumbnail((w, w * 4), Image.LANCZOS)
    if img.mode != "RGB":
        img = img.convert("RGB")
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=80)
    return Response(content=buf.getvalue(), media_type="image/jpeg")


# ─── Visual reviewer state ────────────────────────────────────────────────
# State endpoint returns the sidecar batch.json verbatim if present, else
# synthesises a best-effort view from folder contents. The latter exists so
# legacy batches (processed before the sidecar landed) still render in the
# reviewer — a re-process will replace the synthetic view with the real
# sidecar. `has_sidecar` lets the UI flag the difference.

@app.get("/api/batches/{name}/state")
def batch_state(name: str) -> dict:
    folder = _resolve_batch(name)
    meta = read_meta(folder)
    if meta is not None:
        return {"has_sidecar": True, **meta.model_dump()}
    return _synthesise_state(name, folder)


def _synthesise_state(name: str, folder: Path) -> dict:
    """Build a minimal state view for a batch with no sidecar.

    Images are listed with status inferred from the subfolder they sit in.
    No detection metrics — those would need a re-run. Verdicts default to
    None. Stats are None: there's no group reference without re-running.
    """
    images: list[dict] = []
    for p in sorted(folder.iterdir(), key=lambda x: x.name.lower()):
        if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
            # Original at top-level — status unknown until processed.
            images.append(ImageRow(
                name=p.name,
                status="unprocessed",
                occupied_ratio=0.0,
                confidence=0.0,
                bg_is_white=False,
                bg_purity=0.0,
                output_subfolder=None,
                output_filename=None,
            ).model_dump())
    for sub, status in (("processed", "within-tolerance"), ("review", "review"), ("skipped", "skipped-unknown")):
        sub_dir = folder / sub
        if not sub_dir.is_dir():
            continue
        for p in sorted(sub_dir.iterdir(), key=lambda x: x.name.lower()):
            if p.is_file() and p.suffix.lower() in IMAGE_EXTS:
                # If a top-level original with the same name already exists in
                # `images`, upgrade its status from "unprocessed" instead of
                # adding a duplicate row.
                existing = next((r for r in images if r["name"] == p.name and r["status"] == "unprocessed"), None)
                if existing:
                    existing["status"] = status
                    existing["output_subfolder"] = sub
                    existing["output_filename"] = p.name
                else:
                    images.append(ImageRow(
                        name=p.name,
                        status=status,
                        occupied_ratio=0.0,
                        confidence=0.0,
                        bg_is_white=False,
                        bg_purity=0.0,
                        output_subfolder=sub,
                        output_filename=p.name,
                    ).model_dump())
    return {
        "has_sidecar": False,
        "schema_version": 1,
        "batch_name": name,
        "last_run_timestamp": None,
        "last_run_config": None,
        "stats": None,
        "images": images,
    }


# ─── File-motion helpers for verdict-driven moves ─────────────────────────
# Reject moves the file to skipped/; un-reject restores it to where Picasso
# originally put it (read from the row's `status` field). Centralised here
# so both single-image and bulk verdicts behave identically.

def _current_output_subfolder(folder: Path, filename: str) -> str | None:
    """Where does this image's OUTPUT copy currently live?

    Top-level originals are NOT output copies — they're preserved by
    Picasso's pipeline and verdict-driven moves never touch them. We only
    look at processed/, review/, skipped/.
    """
    for sub in _SUBFOLDERS:
        if (folder / sub / filename).is_file():
            return sub
    return None


def _infer_status_from_location(folder: Path, filename: str) -> str:
    """For legacy batches with no sidecar row, guess what status Picasso
    would have stamped — used to synthesise the `status` field on the
    first verdict so a later un-reject knows where to put the file back.

    Prefers the output-copy location (most informative); falls back to
    "unprocessed" if there's only a top-level original, or "unknown" if
    the file isn't in the batch at all.
    """
    sub = _current_output_subfolder(folder, filename)
    if sub == "processed": return "within-tolerance"
    if sub == "review":    return "review"
    if sub == "skipped":   return "skipped-unknown"
    if (folder / filename).is_file():
        return "unprocessed"
    return "unknown"


def _move_between_subfolders(folder: Path, filename: str, target_sub: str) -> bool:
    """Move the OUTPUT copy of this image into `target_sub`.

    No-op if the file is already there, or if there's no output copy yet
    (i.e., the image only exists at top-level — verdicts don't relocate
    originals). Returns True if a move actually happened.
    """
    src_sub = _current_output_subfolder(folder, filename)
    if src_sub is None or src_sub == target_sub:
        return False
    src = folder / src_sub / filename
    dst = folder / target_sub / filename
    dst.parent.mkdir(parents=True, exist_ok=True)
    if dst.exists():
        # Defensive: clobber any stale destination so the invariant holds.
        try: dst.unlink()
        except OSError: pass
    src.replace(dst)
    return True


def _apply_verdict_motion(
    folder: Path,
    filename: str,
    new_decision: str,
    prior_decision: str | None,
    prior_status: str | None,
) -> str | None:
    """Effect file motion based on the verdict transition. Returns the new
    output_subfolder (or None if no motion happened)."""
    if new_decision == "rejected":
        if _move_between_subfolders(folder, filename, "skipped"):
            return "skipped"
        # Already in skipped/ or only at top-level — record verdict, no
        # location change.
        return None
    if new_decision == "accepted" and prior_decision == "rejected":
        # Un-reject: restore from the sidecar's `status` field.
        target = status_subfolder(prior_status or "")
        if target and target != "skipped":
            if _move_between_subfolders(folder, filename, target):
                return target
    # accept-on-non-rejected, rerun (handled separately), no-op
    return None


# ─── Reviewer verdict + single-image re-run ───────────────────────────────
# One endpoint, three decisions:
#   accepted: record only — the image is good as-is
#   rejected: record only — the image won't ship downstream (filter later)
#   rerun:    re-process this image with user overrides (ratio / canvas /
#             bbox / centroid), write to processed/, clean up stale copies
#             in review/ or skipped/, update sidecar status accordingly.

class VerdictRequest(BaseModel):
    decision: Literal["accepted", "rejected", "rerun"]
    # Re-run overrides — meaningful only when decision == "rerun"
    target_ratio: float | None = Field(default=None, ge=0.0, le=1.0)
    canvas_size: tuple[int, int] | None = None
    bbox: tuple[int, int, int, int] | None = None
    centroid: tuple[float, float] | None = None
    note: str = ""


@app.post("/api/batches/{name}/verdict/{filename}")
async def post_verdict(name: str, filename: str, body: VerdictRequest) -> dict:
    folder = _resolve_batch(name)
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(400, "invalid filename")

    # Validate filename actually belongs to this batch (top-level original or
    # any output subfolder). Without this, a verdict on a typo'd name would
    # silently create a stub row in the sidecar.
    in_batch = (folder / filename).is_file() or any(
        (folder / sub / filename).is_file() for sub in _SUBFOLDERS
    )
    if not in_batch:
        raise HTTPException(404, f"image not in batch: {filename}")

    if body.decision != "rerun":
        # accept/reject path: may move the file (reject → skipped/; accept on
        # a previously-rejected image → restore). Verdict + any location
        # change persisted in one sidecar write.
        meta = read_meta(folder)
        prior_decision = None
        prior_status = None
        if meta:
            for row in meta.images:
                if row.name == filename:
                    if row.verdict:
                        prior_decision = row.verdict.decision
                    prior_status = row.status
                    break
        # First-touch on a legacy batch (no row in sidecar): infer the
        # status from the file's current location so a later un-reject
        # has somewhere to restore to.
        if prior_status is None:
            prior_status = _infer_status_from_location(folder, filename)

        new_subfolder = _apply_verdict_motion(
            folder, filename, body.decision, prior_decision, prior_status
        )

        verdict = ImageVerdict(
            decision=body.decision,
            note=body.note,
            timestamp=now_iso(),
        )
        meta = meta or BatchMeta(batch_name=name)
        found = False
        for row in meta.images:
            if row.name == filename:
                row.verdict = verdict
                if new_subfolder is not None:
                    row.output_subfolder = new_subfolder
                    row.output_filename = filename
                found = True
                break
        if not found:
            meta.images.append(ImageRow(
                name=filename,
                status=prior_status,
                occupied_ratio=0.0,
                confidence=0.0,
                bg_is_white=False,
                bg_purity=0.0,
                output_subfolder=new_subfolder,
                output_filename=filename if new_subfolder else None,
                verdict=verdict,
            ))
        write_meta(folder, meta)
        return {
            "ok": True,
            "decision": body.decision,
            "moved_to": new_subfolder,
        }

    # Re-run path: original at top-level → fresh detect → apply overrides →
    # normalize → write to processed/. Engine work goes to a thread because
    # Pillow holds the GIL during decode and a 12 MP source can stall the
    # event loop otherwise.
    src_path = folder / filename
    if not src_path.is_file():
        raise HTTPException(404, f"source image not at top level: {filename}")

    cfg = load_config(CONFIG_PATH)
    canvas_size = tuple(body.canvas_size) if body.canvas_size else cfg.output_canvas

    # target_ratio resolution: explicit > sidecar's group target > config default > 0.65 fallback.
    target_ratio = body.target_ratio
    if target_ratio is None:
        existing = read_meta(folder)
        if existing and existing.stats:
            target_ratio = float(existing.stats.target_ratio)
        elif isinstance(cfg.target_ratio, (int, float)):
            target_ratio = float(cfg.target_ratio)
        else:
            target_ratio = 0.65

    def _render() -> tuple[bytes, dict]:
        try:
            img = Image.open(src_path)
            img.load()
        except Exception as e:
            raise HTTPException(400, f"cannot read image: {e}")
        det = detect_product(src_path, img, bg_threshold=cfg.bg_threshold)

        # bbox override: must be in-bounds with positive area.
        if body.bbox is not None:
            l, t, r, b = body.bbox
            if not (0 <= l < r <= det.width and 0 <= t < b <= det.height):
                raise HTTPException(422, "bbox out of bounds")
            det.bbox = (l, t, r, b)
            det.occupied_ratio = ((r - l) * (b - t)) / (det.width * det.height)
            if body.centroid is None:
                # Recenter on the new bbox's geometric center.
                det.centroid = ((l + r) / 2.0, (t + b) / 2.0)

        # centroid override: a single click in the original sets where the
        # product lands on the canvas. Forces the mask-centroid path so
        # `det.centroid` is honoured regardless of the global recenter pref.
        recenter_on_mask = cfg.recenter
        if body.centroid is not None:
            cx, cy = body.centroid
            if not (0 <= cx <= det.width and 0 <= cy <= det.height):
                raise HTTPException(422, "centroid out of bounds")
            det.centroid = (float(cx), float(cy))
            recenter_on_mask = True
        # Same logic when only bbox was overridden — we want the recomputed
        # bbox-center to actually be used as the anchor.
        if body.bbox is not None and body.centroid is None:
            recenter_on_mask = True

        normalized = normalize_to_canvas(
            det,
            target_ratio=target_ratio,
            canvas_size=canvas_size,
            padding_pct=cfg.padding_pct,
            max_upscale=cfg.max_upscale,
            recenter_on_mask_centroid=recenter_on_mask,
        )
        out = BytesIO()
        normalized.save(out, format="JPEG", quality=92)
        return out.getvalue(), {
            "occupied_ratio": det.occupied_ratio,
            "confidence": det.confidence,
            "bg_purity": det.bg.purity,
            "bg_is_white": det.bg.is_white,
        }

    out_bytes, det_summary = await asyncio.to_thread(_render)

    # Write to processed/ (overwrites any existing processed copy).
    out_path = folder / "processed" / filename
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(out_bytes)

    # Clean up stale review/ or skipped/ copies — invariant: an image lives
    # in at most one output subfolder.
    for sub in ("review", "skipped"):
        stale = folder / sub / filename
        if stale.exists():
            try:
                stale.unlink()
            except OSError:
                pass  # best-effort; sidecar is the source of truth anyway

    # Sidecar update: row's status & output now reflect the re-run; verdict
    # captures the user's overrides for traceability.
    meta = read_meta(folder) or BatchMeta(batch_name=name)
    verdict = ImageVerdict(
        decision="rerun",
        target_ratio=target_ratio,
        bbox=body.bbox,
        centroid=body.centroid,
        canvas_size=list(canvas_size),
        note=body.note,
        timestamp=now_iso(),
    )
    found = False
    for row in meta.images:
        if row.name == filename:
            row.status = "within-tolerance"
            row.output_subfolder = "processed"
            row.output_filename = filename
            row.occupied_ratio = det_summary["occupied_ratio"]
            row.confidence = det_summary["confidence"]
            row.bg_purity = det_summary["bg_purity"]
            row.bg_is_white = det_summary["bg_is_white"]
            row.verdict = verdict
            found = True
            break
    if not found:
        meta.images.append(ImageRow(
            name=filename,
            status="within-tolerance",
            occupied_ratio=det_summary["occupied_ratio"],
            confidence=det_summary["confidence"],
            bg_is_white=det_summary["bg_is_white"],
            bg_purity=det_summary["bg_purity"],
            output_subfolder="processed",
            output_filename=filename,
            verdict=verdict,
        ))
    write_meta(folder, meta)

    return {
        "ok": True,
        "decision": "rerun",
        "output_path": "processed/" + filename,
        "occupied_ratio": det_summary["occupied_ratio"],
        # b64 lets the UI swap the preview without a second round-trip.
        "image_b64": base64.b64encode(out_bytes).decode("ascii"),
    }


@app.delete("/api/batches/{name}/verdict/{filename}")
def delete_verdict(name: str, filename: str) -> dict:
    """Clear an image's verdict and undo any verdict-driven file motion.

    - Was rejected → restore file from skipped/ to its original location
      (read from `status`). The image returns to the No-verdict queue.
    - Was accepted → metadata-only clear; file stays where it was.
    - Was rerun → only the verdict tag is cleared. The rerun's output
      remains in processed/ because the original processed file was
      overwritten and is unrecoverable; a fresh CLI run on the batch
      would replace it.
    """
    folder = _resolve_batch(name)
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(400, "invalid filename")
    in_batch = (folder / filename).is_file() or any(
        (folder / sub / filename).is_file() for sub in _SUBFOLDERS
    )
    if not in_batch:
        raise HTTPException(404, f"image not in batch: {filename}")

    meta = read_meta(folder)
    if meta is None:
        return {"ok": True, "had_verdict": False, "moved_to": None}

    target_row = next((r for r in meta.images if r.name == filename), None)
    if target_row is None or target_row.verdict is None:
        return {"ok": True, "had_verdict": False, "moved_to": None}

    prior_decision = target_row.verdict.decision
    new_subfolder: str | None = None
    if prior_decision == "rejected":
        # Symmetric with un-reject (accept-after-reject): restore from
        # `status` to the original output subfolder.
        target = status_subfolder(target_row.status or "")
        if target and target != "skipped":
            if _move_between_subfolders(folder, filename, target):
                new_subfolder = target

    target_row.verdict = None
    if new_subfolder is not None:
        target_row.output_subfolder = new_subfolder
        target_row.output_filename = filename
    write_meta(folder, meta)
    return {"ok": True, "had_verdict": True, "moved_to": new_subfolder}


class BulkVerdictRequest(BaseModel):
    decision: Literal["accepted", "rejected"]
    filenames: list[str] = Field(default_factory=list, min_length=1, max_length=5000)
    note: str = ""


@app.post("/api/batches/{name}/verdict-bulk")
def post_verdict_bulk(name: str, body: BulkVerdictRequest) -> dict:
    """Apply the same accept/reject verdict to many images in one round-trip.

    Re-run isn't supported here — each re-run would need its own bbox /
    centroid / target-ratio, which doesn't generalise. Bulk is only useful
    when the verdict itself doesn't carry per-image overrides.

    Reject moves files into skipped/; un-accept (accept on a previously-
    rejected image) restores from the row's status field. All file motion
    happens before the sidecar write so a crash mid-bulk leaves disk and
    sidecar consistent (writes are atomic; moves are too on Win+POSIX).
    """
    folder = _resolve_batch(name)

    meta = read_meta(folder)
    prior_by_name: dict[str, tuple[str | None, str | None]] = {}
    if meta:
        for row in meta.images:
            prior_by_name[row.name] = (
                row.verdict.decision if row.verdict else None,
                row.status,
            )

    # Validate, then perform any file motion. Accumulate per-file results
    # so the sidecar update at the end is one write.
    motions: list[tuple[str, str | None, str]] = []  # (filename, new_subfolder, captured_status)
    errors: list[dict] = []
    ts = now_iso()
    for fname in body.filenames:
        if "/" in fname or "\\" in fname or ".." in fname:
            errors.append({"filename": fname, "error": "invalid filename"})
            continue
        in_batch = (folder / fname).is_file() or any(
            (folder / sub / fname).is_file() for sub in _SUBFOLDERS
        )
        if not in_batch:
            errors.append({"filename": fname, "error": "not in batch"})
            continue
        prior_decision, prior_status = prior_by_name.get(fname, (None, None))
        if prior_status is None:
            prior_status = _infer_status_from_location(folder, fname)
        new_subfolder = _apply_verdict_motion(
            folder, fname, body.decision, prior_decision, prior_status
        )
        motions.append((fname, new_subfolder, prior_status))

    if motions:
        meta = meta or BatchMeta(batch_name=name)
        by_name = {row.name: row for row in meta.images}
        for fname, new_subfolder, prior_status in motions:
            verdict = ImageVerdict(
                decision=body.decision,
                note=body.note,
                timestamp=ts,
            )
            if fname in by_name:
                by_name[fname].verdict = verdict
                if new_subfolder is not None:
                    by_name[fname].output_subfolder = new_subfolder
                    by_name[fname].output_filename = fname
            else:
                meta.images.append(ImageRow(
                    name=fname,
                    status=prior_status,
                    occupied_ratio=0.0,
                    confidence=0.0,
                    bg_is_white=False,
                    bg_purity=0.0,
                    output_subfolder=new_subfolder,
                    output_filename=fname if new_subfolder else None,
                    verdict=verdict,
                ))
        write_meta(folder, meta)
    return {"ok": True, "applied": len(motions), "errors": errors}


class ApplyPresetRequest(BaseModel):
    target_ratio: float | Literal["auto"] | None = None
    canvas_size: tuple[int, int] | None = None


@app.post("/api/batches/{name}/apply-preset")
def apply_preset(name: str, body: ApplyPresetRequest) -> dict:
    """Write target_ratio and/or canvas_size to the batch's folder.yaml.

    `_resolve_config` in cli.py merges folder.yaml on top of imgproc.yaml at
    process time, so the next run on this batch uses these values without any
    other plumbing. Other folder.yaml fields are preserved.
    """
    folder = _resolve_batch(name)
    folder_yaml = folder / "folder.yaml"
    data: dict = {}
    if folder_yaml.exists():
        data = yaml.safe_load(folder_yaml.read_text(encoding="utf-8")) or {}

    if body.target_ratio is not None:
        data["target_ratio"] = body.target_ratio
    if body.canvas_size is not None:
        data["output_canvas"] = list(body.canvas_size)

    # Round-trip through the same Pydantic schema the CLI uses, so an invalid
    # combo (e.g., target_ratio = 1.5) fails here rather than at process time.
    base = Config().model_dump()
    base.update(data)
    try:
        Config(**base)
    except Exception as e:
        raise HTTPException(422, f"invalid preset: {e}")

    folder_yaml.write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    return {"ok": True, "folder_yaml": str(folder_yaml), "saved": data}


# ─── Sheet check (xlsx linter) ────────────────────────────────────────────
# Read-only flagging of common spreadsheet errors before Alida hands an
# xlsx off downstream. Suppressions are persisted next to the xlsx file
# itself (not in the project) so they travel with the workbook.

def _resolve_xlsx_path(raw: str) -> Path:
    """Validate + canonicalise a user-supplied xlsx path.

    The path can sit anywhere on Alida's filesystem (we don't enforce a
    chroot — `import_folder` doesn't either, and this is a single-user
    local UI). What we do enforce: it must exist, be a file, end in
    `.xlsx`, and not be a temporary `~$lockfile`."""
    s = (raw or "").strip().strip('"').strip("'")
    if not s:
        raise HTTPException(400, "xlsx path is required")
    p = Path(s).expanduser()
    if not p.exists() or not p.is_file():
        raise HTTPException(400, f"file not found: {p}")
    if p.suffix.lower() != ".xlsx":
        raise HTTPException(400, "file must have a .xlsx extension")
    if p.name.startswith("~$"):
        raise HTTPException(400, "this looks like an Excel lockfile (~$); pick the real file")
    return p


@app.get("/api/sheetcheck/source-files")
def list_source_xlsx() -> dict:
    """Enumerate .xlsx files under source/ for the picker. Mirrors
    `list_source_folders` — same shape, different filter."""
    source_root = (_PROJECT_ROOT / "source").resolve()
    if not source_root.is_dir():
        return {"root": str(source_root), "exists": False, "files": []}
    files = []
    for p in sorted(source_root.rglob("*.xlsx")):
        if p.name.startswith("~$"):
            continue
        try:
            rel = p.relative_to(source_root)
        except ValueError:
            continue
        files.append({
            "path": str(p),
            "relative": str(rel),
            "size_kb": round(p.stat().st_size / 1024, 1),
        })
    return {"root": str(source_root), "exists": True, "files": files}


class SheetcheckRunRequest(BaseModel):
    xlsx_path: str


@app.post("/api/sheetcheck/run")
def sheetcheck_run(body: SheetcheckRunRequest) -> dict:
    """Parse the xlsx, run rules, return findings split into visible /
    suppressed (per the sidecar). Read-only.
    """
    xlsx_path = _resolve_xlsx_path(body.xlsx_path)
    try:
        parsed = parse_sheet(xlsx_path)
    except Exception as e:
        # parse_sheet wraps zip / corruption errors in RuntimeError —
        # surface as 422 so the UI can show a friendly message.
        raise HTTPException(422, f"could not read xlsx: {e}")
    suffixes = load_suffixes(SUFFIXES_PATH)
    findings = run_rules(parsed, suffixes)
    sup = read_suppressions(xlsx_path)
    visible, suppressed = apply_suppressions(findings, sup)
    return {
        "xlsx_path": str(xlsx_path),
        "sheet_name": parsed.sheet_name,
        "header_row": parsed.header_row,
        "columns": parsed.columns,
        "n_variants": len(parsed.variants),
        "n_images": sum(parsed.images_by_row.values()),
        "parse_warnings": parsed.parse_warnings,
        "suffix_count": len(suffixes.entries),
        "findings": [f.to_dict() for f in visible],
        "suppressed": [f.to_dict() for f in suppressed],
        "muted_findings": sorted(sup.muted_findings),
        "muted_rules": sorted(sup.muted_rules),
        "suppression_sidecar": str(suppression_path(xlsx_path)),
        # `writable` tells the UI whether to surface "Apply" buttons —
        # batch-local xlsx files OR scratch working copies are editable.
        "writable": _is_writable_xlsx(xlsx_path),
        # `is_scratch` lets the UI swap "Make editable copy" for
        # "Save back to source" once a working copy has been opened.
        "is_scratch": _is_inside_scratch(xlsx_path),
        # When the loaded path is in scratch, surface the original
        # source so save-back can target it without the user re-typing.
        "scratch_origin": (
            _read_scratch_origin(xlsx_path) if _is_inside_scratch(xlsx_path) else None
        ),
    }


class SheetcheckSuppressRequest(BaseModel):
    xlsx_path: str
    target: Literal["finding", "rule"]
    key: str                  # suppression_key for "finding"; rule id for "rule"
    action: Literal["mute", "unmute"]


def _is_inside_batches(p: Path) -> bool:
    """True if `p` resolves to a path inside BATCHES_ROOT."""
    try:
        p.resolve().relative_to(BATCHES_ROOT)
        return True
    except ValueError:
        return False


def _is_inside_scratch(p: Path) -> bool:
    """True if `p` resolves to a path inside SCRATCH_ROOT."""
    try:
        p.resolve().relative_to(SCRATCH_ROOT)
        return True
    except ValueError:
        return False


def _is_writable_xlsx(p: Path) -> bool:
    """Whether Sheet check is allowed to write to this path.

    Two writable locations:
      - inside `batches/` — the canonical pre-Pegasus workspace.
      - inside `scratch/sheetcheck/` — standalone working copies for
        the "Make editable copy" flow on xlsx files that haven't been
        imported into a batch yet.

    Source xlsx in `source/` (or anywhere else on disk) stays
    read-only by policy. This guard is what enforces it.
    """
    return _is_inside_batches(p) or _is_inside_scratch(p)


# Sidecar tracking the original source path of a scratch copy. Lets
# "Save back to source" know where to write without the user re-typing.
_SCRATCH_SIDECAR_SUFFIX = ".picasso-source.json"


def _scratch_sidecar(scratch_xlsx: Path) -> Path:
    return scratch_xlsx.with_name(scratch_xlsx.name + _SCRATCH_SIDECAR_SUFFIX)


def _read_scratch_origin(scratch_xlsx: Path) -> dict | None:
    p = _scratch_sidecar(scratch_xlsx)
    if not p.is_file():
        return None
    try:
        import json
        data = json.loads(p.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None
    return data if isinstance(data, dict) else None


def _write_scratch_origin(scratch_xlsx: Path, source_path: Path) -> None:
    import json
    payload = {
        "source_path": str(source_path),
        "scratch_created_at": now_iso(),
        "source_mtime_at_copy": source_path.stat().st_mtime,
    }
    _scratch_sidecar(scratch_xlsx).write_text(
        json.dumps(payload, indent=2), encoding="utf-8"
    )


class SheetcheckApplyFixRequest(BaseModel):
    xlsx_path: str
    row: int                  # 1-indexed
    column: int               # 1-indexed
    value: str
    expected_old: str | None = None  # if provided, refuse the fix when the cell already differs


@app.post("/api/sheetcheck/apply-fix")
def sheetcheck_apply_fix(body: SheetcheckApplyFixRequest) -> dict:
    """Write a single cell value into the xlsx and save.

    Only allowed for xlsx files inside `batches/` — the master copy in
    `source/` is read-only by policy. Returns the updated cell so the
    UI can confirm the write went through.

    `expected_old` is an optional sanity check: if the user sees the
    finding for "RED" but the cell now says something else (file
    edited externally between linter runs), refuse to clobber it. The
    UI surfaces that as "the cell changed; re-run the check".
    """
    xlsx_path = _resolve_xlsx_path(body.xlsx_path)
    if not _is_writable_xlsx(xlsx_path):
        raise HTTPException(
            403,
            "fixes can only be applied to a batch's xlsx or a scratch working copy; "
            "the source xlsx is read-only",
        )
    if body.row < 1 or body.column < 1:
        raise HTTPException(400, "row and column must be 1-indexed positive integers")

    from openpyxl import load_workbook
    try:
        wb = load_workbook(xlsx_path)  # writable mode
    except Exception as e:
        raise HTTPException(422, f"could not open {xlsx_path.name}: {e}")
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    cell = ws.cell(row=body.row, column=body.column)
    if body.expected_old is not None:
        current = "" if cell.value is None else str(cell.value).strip()
        if current.upper() != body.expected_old.strip().upper():
            wb.close()
            raise HTTPException(
                409,
                f"cell already changed (now \"{current}\", expected \"{body.expected_old}\") — re-run check",
            )
    old_value = cell.value
    cell.value = body.value
    try:
        wb.save(xlsx_path)
    except OSError as e:
        wb.close()
        raise HTTPException(500, f"save failed: {e}")
    wb.close()
    return {
        "ok": True,
        "row": body.row,
        "column": body.column,
        "old_value": old_value,
        "new_value": body.value,
    }


class SheetcheckEditCopyRequest(BaseModel):
    xlsx_path: str
    overwrite: bool = False


@app.post("/api/sheetcheck/edit-copy")
def sheetcheck_edit_copy(body: SheetcheckEditCopyRequest) -> dict:
    """Make an editable scratch copy of a (read-only) source xlsx.

    Lets Alida fix typos in a master spreadsheet without first
    creating a Picasso batch. The scratch copy lands in
    `scratch/sheetcheck/{name}.xlsx` and is what the UI loads after
    this endpoint returns; Apply buttons unlock automatically because
    `_is_writable_xlsx` covers the scratch root.

    Refuses (409) when a same-named scratch copy already exists, to
    protect work-in-progress fixes. Pass `overwrite=true` to clobber
    intentionally.
    """
    src = _resolve_xlsx_path(body.xlsx_path)
    if _is_inside_scratch(src):
        raise HTTPException(400, "source path is already inside scratch/")

    SCRATCH_ROOT.mkdir(parents=True, exist_ok=True)
    dest = SCRATCH_ROOT / src.name
    if dest.exists() and not body.overwrite:
        existing_origin = _read_scratch_origin(dest) or {}
        prior_src = existing_origin.get("source_path", "(unknown)")
        raise HTTPException(
            409,
            f"scratch already has {dest.name} (from \"{prior_src}\"); "
            f"finish save-back first or pass overwrite=true",
        )

    try:
        shutil.copy2(src, dest)
        _write_scratch_origin(dest, src)
    except OSError as e:
        # Clean partial state so a retry doesn't trip the overwrite guard.
        try: dest.unlink(missing_ok=True)
        except OSError: pass
        raise HTTPException(500, f"copy failed: {e}")

    return {
        "ok": True,
        "scratch_path": str(dest),
        "source_path": str(src),
    }


class SheetcheckPromoteRequest(BaseModel):
    scratch_path: str
    # Optional override; defaults to the source path saved at copy
    # time. Useful only when the user moved the master xlsx between
    # making the working copy and saving back.
    source_path: str | None = None
    # Bypass the mtime drift check. The UI surfaces a confirm dialog
    # before passing this; we don't blanket-allow.
    force: bool = False


@app.post("/api/sheetcheck/promote-to-source")
def sheetcheck_promote_to_source(body: SheetcheckPromoteRequest) -> dict:
    """Copy the scratch working copy back over the original source.

    Drift guard: if the source's mtime has advanced since the scratch
    copy was made, the user almost certainly edited it in Excel
    meanwhile. Refuse with 409 unless `force=true` so save-back never
    silently clobbers Excel edits.

    On success, deletes the scratch copy + its sidecar so the next
    edit-copy starts fresh.
    """
    scratch = _resolve_xlsx_path(body.scratch_path)
    if not _is_inside_scratch(scratch):
        raise HTTPException(400, "scratch_path must be a working copy under scratch/")

    origin = _read_scratch_origin(scratch) or {}
    target_str = body.source_path or origin.get("source_path")
    if not target_str:
        raise HTTPException(
            422,
            "no source_path recorded for this scratch copy; "
            "pass source_path explicitly",
        )
    target = Path(target_str).expanduser()
    if not target.parent.is_dir():
        raise HTTPException(400, f"source folder not found: {target.parent}")
    if target.suffix.lower() != ".xlsx":
        raise HTTPException(400, "source path must end in .xlsx")

    # Drift check.
    if target.exists() and not body.force:
        original_mtime = origin.get("source_mtime_at_copy")
        current_mtime = target.stat().st_mtime
        if (
            isinstance(original_mtime, (int, float))
            and current_mtime > original_mtime + 0.5  # 0.5s tolerance for FS jitter
        ):
            raise HTTPException(
                409,
                f"source xlsx changed on disk since the working copy was made — "
                f"pass force=true to overwrite anyway, or discard the scratch copy",
            )

    try:
        shutil.copy2(scratch, target)
    except OSError as e:
        raise HTTPException(500, f"save-back failed: {e}")

    # Clean up scratch copy + sidecar — the work has been promoted, the
    # working copy has served its purpose.
    try: scratch.unlink()
    except OSError: pass
    try: _scratch_sidecar(scratch).unlink()
    except OSError: pass

    return {
        "ok": True,
        "saved_to": str(target),
        "scratch_cleared": True,
    }


@app.post("/api/sheetcheck/suppress")
def sheetcheck_suppress(body: SheetcheckSuppressRequest) -> dict:
    """Toggle a per-finding or per-rule mute. Persists to the sidecar
    next to the xlsx — failing that (read-only path) we return a soft
    error so the UI can keep the in-memory mute without crashing."""
    xlsx_path = _resolve_xlsx_path(body.xlsx_path)
    sup = read_suppressions(xlsx_path)
    target_set = sup.muted_findings if body.target == "finding" else sup.muted_rules
    if body.action == "mute":
        target_set.add(body.key)
    else:
        target_set.discard(body.key)
    try:
        path = write_suppressions(xlsx_path, sup)
        return {
            "ok": True,
            "sidecar": str(path),
            "muted_findings": sorted(sup.muted_findings),
            "muted_rules": sorted(sup.muted_rules),
        }
    except OSError as e:
        # Don't 500 — just tell the UI the mute didn't persist. Common case
        # is a read-only network share.
        return {
            "ok": False,
            "error": f"could not write {suppression_path(xlsx_path).name}: {e}",
            "muted_findings": sorted(sup.muted_findings),
            "muted_rules": sorted(sup.muted_rules),
        }


# ─── Visual sort (M5) ─────────────────────────────────────────────────────
# Anchor-load + candidate-hash is heavy (~50 ms/image for bbox-cropped
# pHash on a 12 MP source), so this runs as a background job behind a
# poll-able status endpoint. The job's full result (anchor list, ranked
# matches, dupe clusters) is cached server-side keyed by job_id; the
# apply step replays it without recomputing.

_sort_jobs: dict[str, dict] = {}
_sort_jobs_lock = threading.Lock()
_SORT_TOP_K = 6  # candidates per SKU shown in the picker


def _sort_progress(job_id: str, **fields) -> None:
    with _sort_jobs_lock:
        job = _sort_jobs.get(job_id)
        if job is not None:
            job["progress"].update(fields)


class SortRunRequest(BaseModel):
    xlsx_path: str
    batch_name: str
    threshold: int = 10
    loose_threshold: int = 18
    min_margin: int = 4
    dupe_threshold: int = 10


@app.post("/api/sort/run")
def sort_run(body: SortRunRequest) -> dict:
    """Kick off a sort job. Returns a job_id immediately; poll
    `/api/sort/jobs/{id}` for status and result."""
    folder = _resolve_batch(body.batch_name)
    xlsx_path = _resolve_xlsx_path(body.xlsx_path)

    # Snapshot the candidate set up front. Top-level only; we sort the
    # raws Alida photographed, not the resized outputs.
    candidate_paths = sorted(
        p for p in folder.iterdir()
        if p.is_file() and p.suffix.lower() in IMAGE_EXTS
    )
    if not candidate_paths:
        raise HTTPException(400, f"batch {body.batch_name} has no top-level images")

    job_id = uuid.uuid4().hex
    with _sort_jobs_lock:
        _sort_jobs[job_id] = {
            "status": "running",
            "batch_name": body.batch_name,
            "xlsx_path": str(xlsx_path),
            "progress": {
                "phase": "starting",
                "current": 0,
                "total": len(candidate_paths),
            },
            "result": None,
            "error": None,
        }

    def run() -> None:
        try:
            _sort_progress(job_id, phase="anchors", current=0, total=0)
            anchors = load_anchors_from_xlsx(xlsx_path)
            if not anchors:
                raise RuntimeError("no anchor images found in xlsx")

            _sort_progress(job_id, phase="hashing", current=0, total=len(candidate_paths))

            def cand_progress(done: int, total: int) -> None:
                _sort_progress(job_id, phase="hashing", current=done, total=total)

            candidates = hash_candidates(
                candidate_paths, use_bbox_crop=True, progress=cand_progress,
            )

            _sort_progress(job_id, phase="ranking", current=0, total=len(anchors))
            ranked = rank_candidates_per_sku(
                anchors, candidates,
                threshold=body.threshold,
                loose_threshold=body.loose_threshold,
                min_margin=body.min_margin,
                top_k=_SORT_TOP_K,
            )
            _sort_progress(job_id, phase="dupes", current=0, total=0)
            dupes = find_dupe_clusters(candidates, threshold=body.dupe_threshold)

            with _sort_jobs_lock:
                job = _sort_jobs[job_id]
                job["anchors"] = anchors           # Anchor objects (kept for thumb endpoint)
                job["ranked"] = ranked             # dict[str, list[Match]]
                job["dupes"] = dupes
                job["candidate_paths"] = [p.name for p in candidate_paths]
                job["status"] = "done"
                job["progress"] = {
                    "phase": "done",
                    "current": len(candidate_paths),
                    "total": len(candidate_paths),
                }
        except Exception as e:
            with _sort_jobs_lock:
                job = _sort_jobs.get(job_id)
                if job is not None:
                    job["status"] = "error"
                    job["error"] = f"{type(e).__name__}: {e}"

    threading.Thread(target=run, daemon=True).start()
    return {"job_id": job_id}


@app.get("/api/sort/jobs/{job_id}")
def sort_job(job_id: str) -> dict:
    """Lightweight status poll. Returns progress + status; the full
    result (which contains big data) is fetched separately via
    /api/sort/result/{id} once status == 'done'."""
    with _sort_jobs_lock:
        job = _sort_jobs.get(job_id)
        if job is None:
            raise HTTPException(404, "sort job not found")
        return {
            "status": job["status"],
            "batch_name": job["batch_name"],
            "xlsx_path": job["xlsx_path"],
            "progress": dict(job["progress"]),
            "error": job["error"],
        }


@app.get("/api/sort/result/{job_id}")
def sort_result(job_id: str) -> dict:
    """Full result payload — ranked candidates per SKU, dupe clusters,
    candidate-path list. Anchor thumbnails are served separately via
    /api/sort/anchor/{id}/{sku} so the JSON stays compact."""
    with _sort_jobs_lock:
        job = _sort_jobs.get(job_id)
        if job is None:
            raise HTTPException(404, "sort job not found")
        if job["status"] != "done":
            raise HTTPException(409, f"job not done: {job['status']}")
        anchors: list[Anchor] = job["anchors"]
        ranked: dict[str, list[Match]] = job["ranked"]
        dupes: list[DupeCluster] = job["dupes"]
        candidate_paths: list[str] = job["candidate_paths"]
    skus_payload = []
    for a in anchors:
        matches = ranked.get(a.sku, [])
        skus_payload.append({
            "sku": a.sku,
            "row": a.row,
            "tier": matches[0].tier if matches else "weak",
            "matches": [
                {
                    "filename": m.candidate.name,
                    "distance": m.distance,
                    "rank": m.rank,
                }
                for m in matches
            ],
        })
    dupes_payload = [
        {
            "files": [p.name for p in c.paths],
            "distances": list(c.distances),
        }
        for c in dupes
    ]
    return {
        "batch_name": job["batch_name"],
        "xlsx_path": job["xlsx_path"],
        "skus": skus_payload,
        "dupes": dupes_payload,
        "candidates": candidate_paths,
    }


@app.get("/api/sort/anchor/{job_id}/{sku}")
def sort_anchor(job_id: str, sku: str) -> Response:
    """Serve a SKU's anchor thumbnail. Stable while the job lives in
    memory; once the job is reaped the URL 404s."""
    with _sort_jobs_lock:
        job = _sort_jobs.get(job_id)
        if job is None or job["status"] != "done":
            raise HTTPException(404, "sort job not found")
        anchors: list[Anchor] = job["anchors"]
    for a in anchors:
        if a.sku == sku:
            return Response(content=a.image_bytes, media_type="image/jpeg")
    raise HTTPException(404, "sku not in this job")


class SortMapping(BaseModel):
    sku: str
    hero: str | None = None       # filename at batch top-level; None = skip this SKU
    extras: list[str] = Field(default_factory=list)


class SortApplyRequest(BaseModel):
    job_id: str
    mappings: list[SortMapping]
    overwrite: bool = True   # if False, skip SKUs whose target already exists


@app.post("/api/sort/apply")
def sort_apply(body: SortApplyRequest) -> dict:
    """Effect the chosen mappings. For each SKU we copy:

       processed/{hero}      (if it exists) → processed/sorted/{SKU}.ext
       processed/{extra_i}   (if it exists) → processed/sorted/{SKU}-{b,c,…}.ext

    Falling back to the raw top-level file when the resized version is
    absent — that keeps the tool useful before a user has run a resize
    on the batch, while preferring the white-bg render when available.

    Source files are COPIED, never moved: the batch's top-level
    originals + processed/ outputs stay intact so the user can re-sort
    or re-process. processed/sorted/ is the disposable ship-list."""
    with _sort_jobs_lock:
        job = _sort_jobs.get(body.job_id)
        if job is None:
            raise HTTPException(404, "sort job not found")
        if job["status"] != "done":
            raise HTTPException(409, "sort job not done")
        batch_name = job["batch_name"]

    folder = _resolve_batch(batch_name)
    out_dir = folder / "processed" / "sorted"
    out_dir.mkdir(parents=True, exist_ok=True)

    written: list[dict] = []
    skipped: list[dict] = []
    errors: list[dict] = []

    extras_letters = "bcdefghij"  # -b, -c, -d … extras are unusual past 'd' in practice

    for m in body.mappings:
        if not m.sku:
            errors.append({"sku": m.sku, "error": "blank sku"})
            continue
        if not m.hero:
            skipped.append({"sku": m.sku, "reason": "no hero selected"})
            continue
        # Validate filename + presence at top-level (same safety as elsewhere).
        for fname in [m.hero, *m.extras]:
            if "/" in fname or "\\" in fname or ".." in fname:
                errors.append({"sku": m.sku, "error": f"invalid filename: {fname}"})
                fname = None
                break
        else:
            fname = m.hero  # passed the safety check
        if fname is None:
            continue

        def _resolve_source(name: str) -> Path | None:
            """Prefer the resized output if it exists; otherwise fall
            back to the raw original. Returns None if the file is
            missing entirely."""
            processed = folder / "processed" / name
            if processed.is_file():
                return processed
            top = folder / name
            if top.is_file():
                return top
            return None

        # Hero → {SKU}.{ext}
        src = _resolve_source(m.hero)
        if src is None:
            errors.append({"sku": m.sku, "error": f"hero file not in batch: {m.hero}"})
            continue
        dest = out_dir / f"{m.sku}{src.suffix.lower()}"
        if dest.exists() and not body.overwrite:
            skipped.append({"sku": m.sku, "reason": f"{dest.name} already exists"})
            continue
        try:
            shutil.copy2(src, dest)
        except Exception as e:
            errors.append({"sku": m.sku, "error": f"copy failed: {e}"})
            continue
        written.append({"sku": m.sku, "out": dest.name, "from": src.name})

        # Extras → {SKU}-b.{ext}, {SKU}-c.{ext}, …
        for i, extra in enumerate(m.extras):
            if i >= len(extras_letters):
                errors.append({"sku": m.sku, "error": f"too many extras (>{len(extras_letters)})"})
                break
            esrc = _resolve_source(extra)
            if esrc is None:
                errors.append({"sku": m.sku, "error": f"extra file not in batch: {extra}"})
                continue
            edest = out_dir / f"{m.sku}-{extras_letters[i]}{esrc.suffix.lower()}"
            if edest.exists() and not body.overwrite:
                skipped.append({"sku": m.sku, "reason": f"{edest.name} already exists"})
                continue
            try:
                shutil.copy2(esrc, edest)
            except Exception as e:
                errors.append({"sku": m.sku, "error": f"extra copy failed: {e}"})
                continue
            written.append({"sku": m.sku, "out": edest.name, "from": esrc.name})

    return {
        "ok": True,
        "out_dir": str(out_dir),
        "written": written,
        "skipped": skipped,
        "errors": errors,
    }


# ─── Send to Pegasus (M6) ─────────────────────────────────────────────────
# Picasso writes a batch's processed/sorted/ contents into a configurable
# sync folder. Syncthing (or any equivalent file-sync tool) is what
# actually moves the bytes to the warehouse NUC; Picasso never touches
# the network. Pegasus may later drop a `.picasso-ack.json` ack file
# back into the destination, which we surface as `pegasus_received_at`.

ACK_FILENAME = ".picasso-ack.json"


def _resolve_sync_folder(cfg: Config) -> Path | None:
    """Expand the configured sync_folder. Returns None if unset/blank
    (Send is disabled), or if the path is unreachable. Caller surfaces
    "configure a sync folder first" to the user."""
    raw = (cfg.sync_folder or "").strip().strip('"').strip("'")
    if not raw:
        return None
    return Path(raw).expanduser()


def _read_ack(dest_dir: Path) -> str | None:
    """Read `.picasso-ack.json` if Pegasus has written one; return its
    `received_at` timestamp. Schema is intentionally minimal — Pegasus
    just drops `{ "received_at": "<iso>" }`. Anything malformed or
    missing returns None."""
    ack_path = dest_dir / ACK_FILENAME
    if not ack_path.is_file():
        return None
    try:
        import json
        data = json.loads(ack_path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None
    if not isinstance(data, dict):
        return None
    val = data.get("received_at")
    return val if isinstance(val, str) else None


@app.post("/api/batches/{name}/send")
def send_batch(name: str) -> dict:
    """Copy {batch}/processed/sorted/* into {sync_folder}/{batch_name}/.

    Source = sorted/ exclusively. Without sorted/ the user hasn't run
    the visual sort yet, and shipping un-renamed photos to Pegasus is
    almost never what she wants — fail loudly rather than send a mess.

    Output is a flat folder of SKU-named files, plus the ack timestamp
    if one is already present from a prior cycle. Subsequent sends
    overwrite by default — Syncthing handles the cross-machine reconcile.
    """
    folder = _resolve_batch(name)
    sorted_dir = folder / "processed" / "sorted"
    if not sorted_dir.is_dir():
        raise HTTPException(409, "no processed/sorted/ — run the visual sort first")

    files = sorted(p for p in sorted_dir.iterdir() if p.is_file())
    if not files:
        raise HTTPException(409, "processed/sorted/ is empty — nothing to send")

    cfg = load_config(CONFIG_PATH)
    sync_folder = _resolve_sync_folder(cfg)
    if sync_folder is None:
        raise HTTPException(409, "no sync folder configured — set one in Settings")
    try:
        sync_folder.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        raise HTTPException(422, f"could not create sync folder {sync_folder}: {e}")

    dest_dir = sync_folder / name
    dest_dir.mkdir(parents=True, exist_ok=True)

    written = 0
    for src in files:
        try:
            shutil.copy2(src, dest_dir / src.name)
            written += 1
        except OSError as e:
            raise HTTPException(500, f"copy failed at {src.name}: {e}")

    # Carry the batch's attached xlsx alongside the images so Pegasus
    # has both halves of the curated bundle in one place. Failure is
    # non-fatal — the images sent successfully and that's the bigger
    # half of the work; surface the xlsx error in the response so the
    # UI can flag it.
    meta = read_meta(folder) or BatchMeta(batch_name=name)
    xlsx_sent: str | None = None
    xlsx_error: str | None = None
    xlsx_path = _batch_xlsx_path(folder, meta)
    if xlsx_path is not None:
        try:
            shutil.copy2(xlsx_path, dest_dir / xlsx_path.name)
            xlsx_sent = xlsx_path.name
        except OSError as e:
            xlsx_error = f"xlsx copy failed: {e}"

    ack = _read_ack(dest_dir)
    sent_at = now_iso()

    meta.last_sent_at = sent_at
    meta.last_sent_count = written
    meta.last_sent_dest = str(dest_dir)
    if ack:
        meta.pegasus_received_at = ack
    write_meta(folder, meta)

    return {
        "ok": True,
        "sent_at": sent_at,
        "sent_count": written,
        "dest": str(dest_dir),
        "pegasus_received_at": ack,
        "xlsx_sent": xlsx_sent,
        "xlsx_error": xlsx_error,
    }


@app.get("/api/batches/{name}/send-status")
def send_status(name: str) -> dict:
    """Re-read the ack file without sending. Lets the UI refresh
    `pegasus_received_at` after the user thinks the NUC has had time
    to sync + ack."""
    folder = _resolve_batch(name)
    meta = read_meta(folder)
    last_sent_dest = meta.last_sent_dest if meta else None
    pegasus_received_at = None
    if last_sent_dest:
        ack = _read_ack(Path(last_sent_dest))
        if ack:
            pegasus_received_at = ack
            if meta and meta.pegasus_received_at != ack:
                meta.pegasus_received_at = ack
                write_meta(folder, meta)
    return {
        "last_sent_at": meta.last_sent_at if meta else None,
        "last_sent_count": meta.last_sent_count if meta else None,
        "last_sent_dest": last_sent_dest,
        "pegasus_received_at": pegasus_received_at or (meta.pegasus_received_at if meta else None),
    }


# ─── Server entry ─────────────────────────────────────────────────────────

# ─── Updater ──────────────────────────────────────────────────────────────
# Public-repo GitHub Releases lookup. The check endpoint is hit on every
# page load by the UI banner; install fires on the user's click. Frozen-
# build only (no swap path makes sense in a dev `pip install -e .` setup).

@app.get("/api/updates/check")
def updates_check() -> dict:
    """Return current vs latest version + download URL. Errors surface as
    {has_update: false, error: "..."} so the UI silently no-ops on a bad
    network."""
    from .. import __version__  # local import keeps CLI import-light
    info = check_for_update(__version__)
    return {
        "current_version": info.current_version,
        "latest_version": info.latest_version,
        "has_update": info.has_update,
        "download_url": info.download_url,
        "release_notes": info.release_notes,
        "release_url": info.release_url,
        "error": info.error,
        "is_frozen": bool(getattr(sys, "frozen", False)),
    }


@app.post("/api/updates/install")
def updates_install() -> dict:
    """Kick off the swap. Responds to the client first, then schedules a
    self-exit so the detached swap script can take over without a port
    conflict on relaunch."""
    from .. import __version__
    if not getattr(sys, "frozen", False):
        raise HTTPException(503, "in-app updater only works in packaged builds")

    info = check_for_update(__version__)
    if not info.has_update:
        raise HTTPException(409, f"no update: latest is {info.latest_version}")
    if not info.download_url:
        raise HTTPException(404, "no compatible asset in latest release")

    swap_bat = perform_swap(info.download_url)
    # Defer exit so the response makes it back to the browser.
    threading.Timer(1.0, lambda: os._exit(0)).start()
    return {
        "ok": True,
        "from_version": info.current_version,
        "to_version": info.latest_version,
        "swap_script": str(swap_bat),
    }


def _is_port_taken(host: str, port: int) -> bool:
    """Return True if something is already listening on host:port.

    Uses connect-and-check (not bind) so we don't briefly hold the port
    ourselves — avoids a TOCTOU window where uvicorn would race to bind
    the same port a moment later.
    """
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.settimeout(0.5)
    try:
        s.connect((host, port))
        return True
    except (ConnectionRefusedError, socket.timeout, OSError):
        return False
    finally:
        s.close()


def _open_browser_delayed(url: str, delay_seconds: float = 1.5) -> None:
    """Open the user's browser to `url` after a short delay so uvicorn
    has time to bind. Runs in a daemon thread; failures are silent."""
    def _do() -> None:
        try:
            webbrowser.open(url)
        except Exception:
            pass
    t = threading.Timer(delay_seconds, _do)
    t.daemon = True
    t.start()


def run_server(host: str = "127.0.0.1", port: int = 8765) -> None:
    """Entry point for `imgproc-ui` and the packaged `picasso.exe`.

    Single-instance: if the port is already bound (Picasso already
    running), open the browser to the existing instance and exit instead
    of crashing with a bind error. On a fresh start, spawn a delayed
    browser-open so the user lands on the UI without an extra click.
    """
    url = f"http://{host}:{port}/"
    if _is_port_taken(host, port):
        print(f"Picasso is already running at {url} — opening your browser.")
        try:
            webbrowser.open(url)
        except Exception:
            pass
        return

    _open_browser_delayed(url)
    import uvicorn
    uvicorn.run(app, host=host, port=port, log_level="info")
