"""Sort-only ingest: bucket source images into category subfolders using xlsx anchors.

Uses the bbox-cropped pHash matching approach verified in the 2026-04-24 scoping
experiment (see memory/deferred_xlsx_ingest.md). Each xlsx's filename determines
the category, so:

    2026 CHRISTMAS FLOWERS.xlsx   -> flowers/
    2026 CHRISTMAS BELLS.xlsx     -> bells/
    2026 CHRISTMAS PAPER SERVIETTES.xlsx -> serviettes/

Images whose best anchor match exceeds the Hamming threshold stay in place.
When multiple source images match the same SKU, only the closest match is
moved (renamed to SKU.ext); the rest stay in place for manual review.

Dry-run by default; pass --apply to actually move files.
"""

from __future__ import annotations

import io
import shutil
import sys
from pathlib import Path

import click
import imagehash
from openpyxl import load_workbook
from PIL import Image

from ..engine import detect_product

_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}

# Match xlsx filenames (upper-cased) to category folder names.
_CATEGORIES = {
    "FLOWERS": "flowers",
    "BELLS": "bells",
    "SERVIETTES": "serviettes",
}


def _category_for_xlsx(xlsx_path: Path) -> str | None:
    upper = xlsx_path.name.upper()
    for keyword, folder in _CATEGORIES.items():
        if keyword in upper:
            return folder
    return None


def _bbox_crop(img: Image.Image) -> Image.Image:
    """Run the engine's bbox detector and crop; fall back to the full image if the
    bbox is degenerate (blank image, no product detected)."""
    det = detect_product(Path("anchor"), img)
    left, top, right, bottom = det.bbox
    if right - left < 10 or bottom - top < 10:
        return img
    return img.crop(det.bbox)


def _load_anchors(xlsx_dir: Path) -> dict[str, tuple[imagehash.ImageHash, str]]:
    """Walk every .xlsx in `xlsx_dir`, extract embedded images + their SKU codes,
    and return a dict keyed by SKU: (bbox-cropped pHash, category)."""
    anchors: dict[str, tuple[imagehash.ImageHash, str]] = {}
    for xlsx_path in sorted(xlsx_dir.glob("*.xlsx")):
        if xlsx_path.name.startswith("~$"):
            continue
        category = _category_for_xlsx(xlsx_path)
        if not category:
            continue
        try:
            wb = load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            click.echo(f"  skip {xlsx_path.name}: {e}", err=True)
            continue
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        for img in ws._images:
            anchor_row = img.anchor._from.row + 1
            sku: str | None = None
            # SKUs live in column A slightly above the image anchor — walk up.
            for r in range(anchor_row, max(anchor_row - 20, 0), -1):
                v = ws.cell(row=r, column=1).value
                if v and str(v).strip():
                    sku = str(v).strip()
                    break
            if not sku or sku in anchors:
                continue
            try:
                pil = Image.open(io.BytesIO(img._data())).convert("RGB")
                cropped = _bbox_crop(pil)
                anchors[sku] = (imagehash.phash(cropped), category)
            except Exception:
                continue
    return anchors


@click.command()
@click.argument("source", type=click.Path(exists=True, file_okay=False, path_type=Path))
@click.option(
    "--xlsx-dir",
    type=click.Path(exists=True, file_okay=False, path_type=Path),
    help="Directory containing the category xlsx files. Defaults to SOURCE's parent.",
)
@click.option("--threshold", type=int, default=10, show_default=True,
              help="Strict: accept any match at or below this Hamming distance.")
@click.option("--loose-threshold", type=int, default=18, show_default=True,
              help="Loose: also accept up to this distance IF separation from #2 match is large.")
@click.option("--min-margin", type=int, default=4, show_default=True,
              help="Required gap between #1 and #2 match for a loose accept.")
@click.option("--apply/--dry-run", default=False,
              help="Actually move files. Default is dry-run (preview only).")
def main(
    source: Path,
    xlsx_dir: Path | None,
    threshold: int,
    loose_threshold: int,
    min_margin: int,
    apply: bool,
) -> None:
    """Sort hash-named images in SOURCE into category subfolders (flowers/, bells/,
    serviettes/) based on their best xlsx-anchor match.

    Dry-run by default — review the plan before using --apply.
    """
    # Keep Windows cmd/powershell from choking on UTF-8 output (em-dashes etc.).
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    except Exception:
        pass

    xlsx_dir = xlsx_dir or source.parent
    xlsx_files = [p for p in xlsx_dir.glob("*.xlsx") if not p.name.startswith("~$")]
    if not xlsx_files:
        click.echo(f"No .xlsx files found in {xlsx_dir}", err=True)
        sys.exit(1)

    click.echo(f"Loading anchors from {xlsx_dir} ({len(xlsx_files)} xlsx files)...")
    anchors = _load_anchors(xlsx_dir)
    by_cat: dict[str, int] = {}
    for _, cat in anchors.values():
        by_cat[cat] = by_cat.get(cat, 0) + 1
    cat_summary = ", ".join(f"{c}: {n}" for c, n in sorted(by_cat.items()))
    click.echo(f"  {len(anchors)} anchors loaded ({cat_summary})")

    image_paths = sorted(
        p for p in source.iterdir()
        if p.is_file() and p.suffix.lower() in _IMAGE_EXTS
    )
    click.echo(f"Scoring {len(image_paths)} source images...")

    plan: dict[str, list[tuple[Path, str, int]]] = {
        "flowers": [], "bells": [], "serviettes": [], "unmatched": [],
    }
    n_strict = n_margin = 0
    for p in image_paths:
        try:
            img = Image.open(p).convert("RGB")
            cropped = _bbox_crop(img)
            h = imagehash.phash(cropped)
        except Exception as e:
            click.echo(f"  skip {p.name}: {e}", err=True)
            continue
        # Collect all distances so we can check the #1/#2 separation margin.
        dists = sorted((h - ah, sku, cat) for sku, (ah, cat) in anchors.items())
        top1_dist, top1_sku, top1_cat = dists[0]
        top2_dist = dists[1][0] if len(dists) > 1 else 999

        strict = top1_dist <= threshold
        # Loose: further than strict but still plausible, AND clearly separated from
        # the next candidate. The margin guards against noise-floor false positives.
        margin_ok = (
            top1_dist <= loose_threshold
            and (top2_dist - top1_dist) >= min_margin
        )
        if strict:
            plan[top1_cat].append((p, top1_sku, top1_dist))
            n_strict += 1
        elif margin_ok:
            plan[top1_cat].append((p, top1_sku, top1_dist))
            n_margin += 1
        else:
            plan["unmatched"].append((p, top1_sku, top1_dist))

    # Per SKU, the closest-match image is the hero; any additional matches are
    # "extras" (usually lifestyle/packaging shots of the same product) and stay
    # in place, so the category folder ends up with one clean image per SKU.
    heroes_by_cat: dict[str, list[tuple[Path, str, int]]] = {
        "flowers": [], "bells": [], "serviettes": [],
    }
    n_extras_by_cat: dict[str, int] = {"flowers": 0, "bells": 0, "serviettes": 0}
    for cat in ("flowers", "bells", "serviettes"):
        by_sku: dict[str, list[tuple[Path, int]]] = {}
        for p, sku, dist in plan[cat]:
            by_sku.setdefault(sku, []).append((p, dist))
        for sku, entries in by_sku.items():
            entries.sort(key=lambda e: e[1])
            hero_p, hero_dist = entries[0]
            heroes_by_cat[cat].append((hero_p, sku, hero_dist))
            n_extras_by_cat[cat] += len(entries) - 1

    click.echo("\nSort plan:")
    for cat in ("flowers", "bells", "serviettes"):
        extras = n_extras_by_cat[cat]
        extra_note = f" (+{extras} extras left in place)" if extras else ""
        click.echo(f"  {cat:12s}: {len(heroes_by_cat[cat])}{extra_note}")
    click.echo(f"  unmatched   : {len(plan['unmatched'])} (will stay in place)")
    click.echo(f"  (of matches: {n_strict} strict + {n_margin} via separation margin)")

    if not apply:
        click.echo("\n(dry-run — re-run with --apply to actually move files)")
        return

    click.echo("\nMoving files...")
    renames_preview: list[tuple[str, str]] = []
    for cat in ("flowers", "bells", "serviettes"):
        if not heroes_by_cat[cat]:
            continue
        target = source / cat
        target.mkdir(exist_ok=True)
        for p, sku, _ in heroes_by_cat[cat]:
            new_name = f"{sku}{p.suffix.lower()}"
            dest = target / new_name
            # Collision is unlikely now that only one file per SKU moves, but a
            # prior --apply run could have left SKU.ext in place already.
            if dest.exists():
                dest = target / f"{sku}-{p.stem[:8]}{p.suffix.lower()}"
            shutil.move(str(p), str(dest))
            renames_preview.append((p.name, f"{cat}/{dest.name}"))
    # Print the first few renames as a sanity check, then summary counts.
    for old, new in renames_preview[:6]:
        click.echo(f"  {old}  ->  {new}")
    if len(renames_preview) > 6:
        click.echo(f"  ...and {len(renames_preview) - 6} more")
    click.echo("Done.")


if __name__ == "__main__":
    main()
