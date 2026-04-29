"""Shared anchor-load + match logic, used by both the CLI in
`ingest/sort.py` and the visual sort UI in `web/app.py`.

Extracted from the CLI so the web layer can call it without
shelling out (and without owning the bbox/imagehash dependencies twice).

Anchors live in the spreadsheet — every variant SKU has an embedded
product photo. We pHash each anchor (after bbox-cropping to the product)
once, then compare every candidate image's pHash by Hamming distance.

Hash mode is fixed at `phash`: per the v1 plan we explicitly avoid the
imagehash modes that lazy-import scipy (~80 MB). pyproject pins
`imagehash>=4.3` and the spec excludes scipy.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from pathlib import Path

import imagehash
from openpyxl import load_workbook
from PIL import Image

from ..engine import detect_product


# ─── bbox helper ──────────────────────────────────────────────────────

def bbox_crop(img: Image.Image) -> Image.Image:
    """Crop to the detected product bbox, falling back to the full
    image if the detector returns a degenerate region (blank canvas,
    no product). Public so callers can pHash crops consistently with
    how anchors were hashed."""
    det = detect_product(Path("anchor"), img)
    left, top, right, bottom = det.bbox
    if right - left < 10 or bottom - top < 10:
        return img
    return img.crop(det.bbox)


# ─── Anchors ──────────────────────────────────────────────────────────

@dataclass
class Anchor:
    sku: str
    phash: imagehash.ImageHash
    image_bytes: bytes      # JPEG bytes of the anchor (full, not the crop) — for UI thumbnail
    row: int                # 1-indexed anchor row in the xlsx (for tooltips)


def load_anchors_from_xlsx(xlsx_path: Path) -> list[Anchor]:
    """Single-xlsx anchor loader. Iterates embedded images, walks up to
    find the SKU in column A, pHashes the bbox-cropped product.

    Multiple anchors per SKU are deduplicated by keeping the first
    occurrence (matches the CLI's behaviour). Returns a list (ordered
    by appearance in the sheet) so the UI shows SKUs in spreadsheet
    order, not alphabetical.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    anchors: list[Anchor] = []
    seen: set[str] = set()
    for img in getattr(ws, "_images", []):
        try:
            anchor_row = img.anchor._from.row + 1
        except AttributeError:
            continue
        sku: str | None = None
        for r in range(anchor_row, max(anchor_row - 20, 0), -1):
            v = ws.cell(row=r, column=1).value
            if v and str(v).strip():
                sku = str(v).strip()
                break
        if not sku or sku in seen:
            continue
        try:
            data = img._data()
            pil = Image.open(io.BytesIO(data)).convert("RGB")
            cropped = bbox_crop(pil)
            ph = imagehash.phash(cropped)
        except Exception:
            continue
        # Re-encode the original (uncropped) anchor as a small JPEG for the
        # browser. The xlsx-embedded format is sometimes PNG; this gives
        # us a uniform thumbnail.
        thumb = pil.copy()
        thumb.thumbnail((200, 200), Image.LANCZOS)
        buf = io.BytesIO()
        thumb.save(buf, format="JPEG", quality=82)
        anchors.append(Anchor(
            sku=sku, phash=ph, image_bytes=buf.getvalue(), row=anchor_row,
        ))
        seen.add(sku)
    wb.close()
    return anchors


# ─── Candidate hashing ────────────────────────────────────────────────

@dataclass
class Candidate:
    path: Path
    phash: imagehash.ImageHash


def hash_candidates(
    paths: list[Path],
    *,
    use_bbox_crop: bool,
    progress=None,
) -> list[Candidate]:
    """Compute pHash for every path. Bad files (unreadable / decode
    errors) are silently skipped — they show up as missing in the
    candidate list rather than crashing the whole batch.

    `use_bbox_crop` should mirror how the anchors were hashed. For
    anchors we always bbox-crop (their backgrounds are noisy); for
    matching against `processed/` images (already 600×800 white-bg)
    bbox-crop adds little. Keep callers explicit.
    """
    out: list[Candidate] = []
    for i, p in enumerate(paths):
        try:
            img = Image.open(p).convert("RGB")
            if use_bbox_crop:
                img = bbox_crop(img)
            out.append(Candidate(path=p, phash=imagehash.phash(img)))
        except Exception:
            continue
        if progress and (i % 5 == 0 or i + 1 == len(paths)):
            progress(i + 1, len(paths))
    if progress:
        progress(len(paths), len(paths))
    return out


# ─── Match scoring ────────────────────────────────────────────────────

@dataclass
class Match:
    """One (SKU, candidate) pair's distance + tier."""
    sku: str
    candidate: Path
    distance: int
    rank: int               # 0 = best match for this SKU, 1 = second, …
    tier: str               # "strict" | "margin" | "weak"


def rank_candidates_per_sku(
    anchors: list[Anchor],
    candidates: list[Candidate],
    *,
    threshold: int = 10,
    loose_threshold: int = 18,
    min_margin: int = 4,
    top_k: int = 6,
) -> dict[str, list[Match]]:
    """For every SKU return the top-K closest candidates by Hamming
    distance. The first match's tier classifies how confident the
    auto-pick would be:

        strict — distance ≤ `threshold`
        margin — distance ≤ `loose_threshold` AND clearly separated
                 from the runner-up by ≥ `min_margin`
        weak   — anything else (UI shows but doesn't auto-pick)

    Tier is recorded only on the rank-0 entry; ranks 1..K-1 just carry
    distance.
    """
    out: dict[str, list[Match]] = {}
    for a in anchors:
        scored = sorted(
            ((a.phash - c.phash, c) for c in candidates),
            key=lambda t: t[0],
        )[:top_k]
        if not scored:
            out[a.sku] = []
            continue
        top1 = scored[0][0]
        top2 = scored[1][0] if len(scored) > 1 else 999
        if top1 <= threshold:
            tier = "strict"
        elif top1 <= loose_threshold and (top2 - top1) >= min_margin:
            tier = "margin"
        else:
            tier = "weak"
        out[a.sku] = [
            Match(
                sku=a.sku, candidate=c.path, distance=int(d),
                rank=i, tier=(tier if i == 0 else ""),
            )
            for i, (d, c) in enumerate(scored)
        ]
    return out


# ─── Dupe detection ───────────────────────────────────────────────────

@dataclass
class DupeCluster:
    """A set of candidate paths whose pHashes are pairwise within
    `threshold`. Order: representative first (lowest path name), then
    the rest sorted by distance from it."""
    paths: list[Path]
    distances: list[int]    # parallel to paths; representative's distance is 0


def find_dupe_clusters(
    candidates: list[Candidate],
    *,
    threshold: int = 10,
) -> list[DupeCluster]:
    """Greedy single-link clustering: walk candidates in name order;
    each one either joins an existing cluster (distance ≤ threshold to
    its representative) or starts a new one.

    Single-link is sufficient at 200-image scale. Singletons are
    dropped — the caller only cares about lookalikes.
    """
    sorted_cands = sorted(candidates, key=lambda c: c.path.name.lower())
    cluster_reps: list[Candidate] = []
    membership: list[list[tuple[Candidate, int]]] = []  # parallel to cluster_reps
    for c in sorted_cands:
        joined = False
        for i, rep in enumerate(cluster_reps):
            d = int(c.phash - rep.phash)
            if d <= threshold:
                membership[i].append((c, d))
                joined = True
                break
        if not joined:
            cluster_reps.append(c)
            membership.append([(c, 0)])
    clusters: list[DupeCluster] = []
    for members in membership:
        if len(members) < 2:
            continue
        members.sort(key=lambda t: t[1])
        clusters.append(DupeCluster(
            paths=[m[0].path for m in members],
            distances=[m[1] for m in members],
        ))
    return clusters
