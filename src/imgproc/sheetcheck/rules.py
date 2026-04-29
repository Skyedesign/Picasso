"""Parse a pre-buy xlsx and run lint rules against its variant rows.

Variant detection follows Alida's convention: a variant is any row whose
QTY column is non-blank. The rest of the rows hold image anchors,
manufacturer URLs, blank padding, etc. — those are interesting only
inasmuch as they tell us which images belong to which variant.

Rules are deterministic and read-only — no AI, no fuzzy "did you mean"
beyond simple typo heuristics. The acceptance criterion (CLAUDE.md / the
plan) is the textiles MATERIAL stack must NOT trip a false positive, so
DESCRIPTION + MATERIAL are explicitly exempt from the blank-cell rule.
"""

from __future__ import annotations

import re
import time
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

from .suffixes import SuffixDict, SuffixEntry


# Header tokens we look for in the first ~10 rows. Capitalisation varies
# across sheets; matching is case-insensitive.
HEADER_TOKENS = {
    "code": ["CODE", "SKU", "#"],
    "description": ["DESCRIPTION", "DESC"],
    "size": ["SIZE"],
    "material": ["MATERIAL"],
    "colour": ["COLOUR", "COLOR"],
    "qty": ["QTY", "QUANTITY"],
}

# Required-non-blank rule applies to these only (per plan: DESCRIPTION
# and MATERIAL exempt because of multi-component textile layouts).
REQUIRED_COLUMNS: tuple[str, ...] = ("size", "colour")


# ─── Schema ────────────────────────────────────────────────────────────

Severity = Literal["error", "warning", "info"]


@dataclass
class Variant:
    """One variant row's parsed contents + the row span it occupies."""
    sku: str
    row: int             # 1-indexed
    block_start: int     # 1-indexed; first row of this variant's block
    block_end: int       # 1-indexed; INCLUSIVE last row of this block
    cells: dict[str, Any]   # keyed by logical column name (lowercase)


@dataclass
class ParsedSheet:
    xlsx_path: str
    sheet_name: str
    header_row: int
    columns: dict[str, int]       # logical name → 1-indexed column number
    variants: list[Variant] = field(default_factory=list)
    images_by_row: dict[int, int] = field(default_factory=dict)  # row → count
    parse_warnings: list[str] = field(default_factory=list)


@dataclass
class Finding:
    rule: str                 # rule id, e.g. "blank_required_column"
    severity: Severity
    row: int                  # 1-indexed; 0 if not row-specific
    sku: str | None
    message: str              # human-readable summary
    suggestion: str | None = None
    suppression_key: str = ""  # stable key for per-row+rule mutes
    # Optional structured "Apply" payload — populated for findings that
    # have a deterministic single-cell fix (suffix mismatch ⇒ replace
    # COLOUR cell; SKU family break ⇒ replace CODE cell). None means
    # the user has to fix it manually. Schema:
    #   {"row": int, "column": int (1-indexed), "value": str}
    fix: dict[str, Any] | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "rule": self.rule,
            "severity": self.severity,
            "row": self.row,
            "sku": self.sku,
            "message": self.message,
            "suggestion": self.suggestion,
            "suppression_key": self.suppression_key,
            "fix": self.fix,
        }


# ─── Parsing ───────────────────────────────────────────────────────────

def _open_workbook(xlsx_path: Path):
    """`read_only=True, data_only=True` per the plan. Excel can leave the
    file in a partially-zipped state mid-save; one retry handles that
    transient case."""
    last_err: Exception | None = None
    for attempt in range(2):
        try:
            return load_workbook(xlsx_path, read_only=True, data_only=True)
        except zipfile.BadZipFile as e:
            last_err = e
            time.sleep(0.2)
        except Exception as e:
            last_err = e
            break
    raise RuntimeError(f"could not open {xlsx_path.name}: {last_err}")


def _detect_header(ws, max_scan: int = 12) -> tuple[int, dict[str, int]]:
    """Walk the first `max_scan` rows looking for the row that contains
    QTY + at least one of CODE / SKU. Returns (1-indexed row, column map).

    Falls back to row 1 if nothing matches; the caller surfaces a
    parse_warning rather than crashing — the rest of the linter can
    still run on whatever variants we extract.
    """
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        cols: dict[str, int] = {}
        for c_idx, val in enumerate(row, start=1):
            if val is None:
                continue
            v = str(val).strip().upper()
            if not v:
                continue
            for logical, tokens in HEADER_TOKENS.items():
                if logical in cols:
                    continue
                if v in tokens:
                    cols[logical] = c_idx
                    break
        if "qty" in cols and ("code" in cols or 1 in cols.values()):
            return r_idx, cols
    return 1, {}


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _is_variant_row(ws, r: int, qty_col: int, code_col: int) -> bool:
    """A row is a variant iff QTY is non-blank. CODE non-blank without
    QTY is the "section header" row pattern Alida sometimes uses; we
    skip those.
    """
    qty = ws.cell(row=r, column=qty_col).value
    if qty is None:
        return False
    code = ws.cell(row=r, column=code_col).value if code_col else None
    # Header-row look-alikes: ignore the row matching the column-name pattern.
    code_s = _safe_str(code).upper()
    if code_s in {"CODE", "SKU", "#"}:
        return False
    return True


def parse_sheet(xlsx_path: Path) -> ParsedSheet:
    """Open the xlsx and extract variants + image anchors. Read-only;
    nothing is written.

    Falls back gracefully on a missing header — the column map will be
    partial, and downstream rules that rely on missing columns will be
    no-ops rather than raising.
    """
    wb = _open_workbook(xlsx_path)
    # Plan: prefer "Sheet1" if present; otherwise the active sheet.
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    parsed = ParsedSheet(
        xlsx_path=str(xlsx_path),
        sheet_name=ws.title,
        header_row=1,
        columns={},
    )

    # `read_only=True` workbooks don't expose `_images`, and openpyxl is
    # actively hostile to mixing the two modes — so re-open without
    # read_only specifically for image anchors.
    code_col = 1
    try:
        wb2 = load_workbook(xlsx_path, data_only=True)
        ws2 = wb2["Sheet1"] if "Sheet1" in wb2.sheetnames else wb2.active
        for img in getattr(ws2, "_images", []):
            try:
                row_1based = img.anchor._from.row + 1
            except AttributeError:
                continue
            parsed.images_by_row[row_1based] = parsed.images_by_row.get(row_1based, 0) + 1
        wb2.close()
    except Exception as e:
        parsed.parse_warnings.append(f"could not enumerate embedded images: {e}")

    parsed.header_row, parsed.columns = _detect_header(ws)
    if not parsed.columns:
        parsed.parse_warnings.append("no recognisable header row in first 12 rows")
        return parsed

    qty_col = parsed.columns.get("qty")
    code_col = parsed.columns.get("code", 1)

    # Walk rows after the header. We materialise variant rows first, then
    # set each variant's block end to the row just before the next
    # variant (or the last sheet row).
    variant_rows: list[int] = []
    if qty_col:
        for r in range(parsed.header_row + 1, ws.max_row + 1):
            if _is_variant_row(ws, r, qty_col=qty_col, code_col=code_col):
                variant_rows.append(r)

    for i, r in enumerate(variant_rows):
        block_end = (variant_rows[i + 1] - 1) if i + 1 < len(variant_rows) else ws.max_row
        cells: dict[str, Any] = {}
        for logical, col_idx in parsed.columns.items():
            cells[logical] = ws.cell(row=r, column=col_idx).value
        sku = _safe_str(cells.get("code"))
        parsed.variants.append(Variant(
            sku=sku, row=r, block_start=r, block_end=block_end, cells=cells,
        ))
    wb.close()
    return parsed


# ─── Rule helpers ──────────────────────────────────────────────────────

# A "family" is letters + the first contiguous digit run; i.e. the
# numeric-family identifier the variant_gap rule groups by. Suffixes
# (everything after the first dash, or trailing letters after digits)
# are excluded.
_FAMILY_RE = re.compile(r"^([A-Z]+)(\d+)")


def _family_key(sku: str) -> tuple[str, int] | None:
    m = _FAMILY_RE.match(sku.upper())
    if not m:
        return None
    return m.group(1), int(m.group(2))


def _suppression_key(rule: str, row: int, *parts: str) -> str:
    extra = ":" + ":".join(parts) if parts else ""
    return f"{rule}@row{row}{extra}"


# ─── Rules ─────────────────────────────────────────────────────────────
# Each rule is a function that takes the parsed sheet (+ optionally the
# suffix dict) and returns a list of Findings. Adding a rule = adding
# a function + registering it in `RULES` below. No clever framework.

def rule_blank_required_column(parsed: ParsedSheet) -> list[Finding]:
    """SIZE / COLOUR cells must be non-blank on a variant row.

    DESCRIPTION + MATERIAL are exempt: textile rows leave them blank when
    a downstream "components" sub-table breaks the value across multiple
    rows. Per the v1 plan, those columns are deliberately out of scope.
    """
    out: list[Finding] = []
    for v in parsed.variants:
        for col in REQUIRED_COLUMNS:
            if col not in parsed.columns:
                continue  # column absent from this sheet; can't check
            cell = _safe_str(v.cells.get(col))
            if cell:
                continue
            out.append(Finding(
                rule="blank_required_column",
                severity="warning",
                row=v.row,
                sku=v.sku or None,
                message=f"{col.upper()} is blank",
                suggestion=None,
                suppression_key=_suppression_key("blank_required_column", v.row, col),
            ))
    return out


def rule_suffix_column_mismatch(parsed: ParsedSheet, suffixes: SuffixDict) -> list[Finding]:
    """For each variant, take the SKU's trailing suffix and verify the
    cell on the suffix's expected column doesn't contradict it.

    Conservative — only flags when the cell contains a *competing*
    keyword from a different entry mapped to the same column. Unknown
    cell text → no opinion (the dictionary is incomplete by design).
    """
    if not suffixes.entries:
        return []
    out: list[Finding] = []
    for v in parsed.variants:
        if not v.sku:
            continue
        entry: SuffixEntry | None = suffixes.detect(v.sku)
        if entry is None:
            continue
        col_name = entry.column.lower()
        if col_name not in parsed.columns:
            continue
        cell = _safe_str(v.cells.get(col_name))
        if not cell:
            continue  # blank handled by the required-column rule
        if suffixes.cell_matches(entry, cell):
            continue
        competing = suffixes.competing_match(entry, cell)
        if competing is None:
            continue  # ambiguous / not in dict — don't flag
        suggestion = (
            f"SKU suffix -{entry.id} usually means {entry.keywords[0].title()}; "
            f"this row's {entry.column} reads \"{cell}\" "
            f"(matches sibling -{competing.id} = {competing.keywords[0].title()})."
        )
        # Auto-fix: replace the offending cell with the suffix's primary
        # keyword. We use keywords[0] as the canonical form (e.g. "GREEN"
        # for `-G`); the dictionary's first entry is the conventional one.
        target_col = parsed.columns[col_name]
        fix_payload = {
            "row": v.row,
            "column": target_col,
            "value": entry.keywords[0],
        }
        out.append(Finding(
            rule="suffix_column_mismatch",
            severity="warning",
            row=v.row,
            sku=v.sku,
            message=f"{entry.column} \"{cell}\" doesn't match SKU suffix -{entry.id}",
            suggestion=suggestion,
            suppression_key=_suppression_key("suffix_column_mismatch", v.row),
            fix=fix_payload,
        ))
    return out


def rule_missing_image(parsed: ParsedSheet) -> list[Finding]:
    """Each variant block should contain at least one anchored image.

    Anchor row range = [block_start, block_end]. Images embedded outside
    every variant's block are caught by the orphan-image rule.
    """
    out: list[Finding] = []
    for v in parsed.variants:
        anchored = sum(
            count for row, count in parsed.images_by_row.items()
            if v.block_start <= row <= v.block_end
        )
        if anchored == 0:
            out.append(Finding(
                rule="missing_image",
                severity="warning",
                row=v.row,
                sku=v.sku or None,
                message=f"no embedded image found for {v.sku or 'this variant'}",
                suggestion="Drop the product photo into the spreadsheet for this row.",
                suppression_key=_suppression_key("missing_image", v.row),
            ))
    return out


def rule_image_sku_correlation(parsed: ParsedSheet) -> list[Finding]:
    """Sanity-check image placement: every embedded image should anchor
    inside some variant's block. An orphan image (anchored to no
    variant's range) usually means rows shifted around the image, so
    the SKU label and the picture no longer agree.

    Reported once per orphan-image row, not per-variant — too noisy
    otherwise.
    """
    out: list[Finding] = []
    if not parsed.variants:
        return out
    blocks = [(v.block_start, v.block_end, v.sku, v.row) for v in parsed.variants]
    for img_row in sorted(parsed.images_by_row.keys()):
        in_block = any(start <= img_row <= end for start, end, _, _ in blocks)
        if in_block:
            continue
        # Find the closest variant for the suggestion text.
        nearest = min(blocks, key=lambda b: min(abs(img_row - b[0]), abs(img_row - b[1])))
        out.append(Finding(
            rule="image_sku_correlation",
            severity="warning",
            row=img_row,
            sku=None,
            message=f"image anchored at row {img_row} sits outside every variant block",
            suggestion=f"Closest variant: {nearest[2] or '?'} (rows {nearest[0]}–{nearest[1]}).",
            suppression_key=_suppression_key("image_sku_correlation", img_row),
        ))
    return out


_DASHED_SKU_RE = re.compile(r"^([A-Z]+)(\d+)-(.+)$")
_UNDASHED_SKU_RE = re.compile(r"^([A-Z]+)(\d+)$")


def rule_sku_family_break(parsed: ParsedSheet) -> set[int]:
    """SKU typo detector: flag a SKU that lacks the dash convention used
    by a recognisable sibling family.

    Returns the set of variant rows it flagged so other rules
    (variant_gap) can ignore those rows — otherwise an `E0120` typo
    inflates the family-number range and yields a wall of false
    positives.

    Approach:
      1. Build the set of sibling-family prefixes from dashed SKUs:
         `A04-G` ⇒ prefix "A04". A prefix needs ≥ 2 dashed members to
         be considered "established".
      2. For each undashed SKU, walk every possible split `letters +
         digits` ⇒ `<prefix>` + `<tail>`. The first split whose prefix
         matches an established sibling family is the suggested fix.
    """
    out_rows: set[int] = set()
    findings: list[Finding] = []

    sibling_families: dict[str, int] = {}  # prefix → count of dashed siblings
    for v in parsed.variants:
        if not v.sku:
            continue
        m = _DASHED_SKU_RE.match(v.sku.upper())
        if m:
            prefix = m.group(1) + m.group(2)
            sibling_families[prefix] = sibling_families.get(prefix, 0) + 1

    for v in parsed.variants:
        if not v.sku:
            continue
        sku_u = v.sku.upper()
        if _DASHED_SKU_RE.match(sku_u):
            continue  # well-formed
        m = _UNDASHED_SKU_RE.match(sku_u)
        if not m:
            continue  # too unusual to second-guess (e.g., free-text product code)
        letters, digits = m.group(1), m.group(2)
        # Try splits from longest-prefix-first so `A0420` prefers `A042` if
        # `A042-*` exists (unlikely but possible) before falling back to
        # `A04`. Skip splits that produce an empty tail.
        suggestion = None
        for split in range(len(digits) - 1, 0, -1):
            cand_prefix = letters + digits[:split]
            cand_tail = digits[split:]
            if sibling_families.get(cand_prefix, 0) >= 2:
                suggestion = f"{cand_prefix}-{cand_tail}"
                break
        if suggestion is None:
            continue
        out_rows.add(v.row)
        # Auto-fix: replace the SKU code cell (column 1) with the
        # suggested form. Code column is conventionally column 1; if a
        # future sheet places it elsewhere, parsed.columns["code"] is
        # the authoritative source.
        target_col = parsed.columns.get("code", 1)
        fix_payload = {
            "row": v.row,
            "column": target_col,
            "value": suggestion,
        }
        findings.append(Finding(
            rule="sku_family_break",
            severity="warning",
            row=v.row,
            sku=v.sku,
            message=f"SKU \"{v.sku}\" breaks the {suggestion.split('-')[0]}-* convention used by its siblings",
            suggestion=f"Did you mean \"{suggestion}\"?",
            suppression_key=_suppression_key("sku_family_break", v.row),
            fix=fix_payload,
        ))
    parsed._family_break_findings = findings  # type: ignore[attr-defined]
    return out_rows


def rule_variant_gap(parsed: ParsedSheet, exclude_rows: set[int]) -> list[Finding]:
    """Gentle nudge when a numeric family is missing a near-neighbour,
    e.g. A01, A02, A04 with A03 absent.

    Only short gaps (1–3 missing numbers between observed siblings) are
    reported. Long jumps (A04 → A420) almost always reflect a SKU typo
    or a deliberate range jump rather than a missed entry, so flagging
    them buries the real signal. Rows already flagged by
    `sku_family_break` are excluded so a typo doesn't double-count.

    Severity is "info": Alida sometimes retires SKUs deliberately, and
    the rule is meant as a nudge, not a blocker.
    """
    out: list[Finding] = []
    by_prefix: dict[str, list[tuple[int, int]]] = {}  # prefix → [(fam_num, row), …]
    for v in parsed.variants:
        if not v.sku or v.row in exclude_rows:
            continue
        key = _family_key(v.sku)
        if key is None:
            continue
        prefix, fam_num = key
        by_prefix.setdefault(prefix, []).append((fam_num, v.row))

    MAX_GAP_SIZE = 3  # tolerate one or two missing siblings; bigger is probably intentional

    for prefix, members in by_prefix.items():
        # Unique family numbers, with the row of their first occurrence.
        first_row: dict[int, int] = {}
        for fam_num, row in members:
            first_row.setdefault(fam_num, row)
        nums_sorted = sorted(first_row.keys())
        if len(nums_sorted) < 3:
            continue  # too few siblings to draw a confident pattern
        for a, b in zip(nums_sorted, nums_sorted[1:]):
            gap = b - a - 1
            if gap < 1 or gap > MAX_GAP_SIZE:
                continue
            missing = list(range(a + 1, b))
            gap_str = ", ".join(
                f"{prefix}{n:02d}" if n < 100 else f"{prefix}{n}" for n in missing
            )
            out.append(Finding(
                rule="variant_gap",
                severity="info",
                row=first_row[a],
                sku=None,
                message=f"{prefix} family skips {gap_str} (between {prefix}{a:02d} and {prefix}{b:02d})",
                suggestion="Was that intentional? Otherwise a row got missed.",
                suppression_key=_suppression_key("variant_gap", first_row[a], prefix),
            ))
    return out


# Ordered registration → ordered output. Order matters only for stable UI.
RULES = (
    "blank_required_column",
    "suffix_column_mismatch",
    "sku_family_break",
    "image_sku_correlation",
    "missing_image",
    "variant_gap",
)


def run_rules(parsed: ParsedSheet, suffixes: SuffixDict) -> list[Finding]:
    findings: list[Finding] = []
    findings += rule_blank_required_column(parsed)
    findings += rule_suffix_column_mismatch(parsed, suffixes)
    # family-break runs first among the family-aware rules so its
    # excluded-row set can be passed to variant_gap.
    family_break_rows = rule_sku_family_break(parsed)
    findings += getattr(parsed, "_family_break_findings", [])
    findings += rule_image_sku_correlation(parsed)
    findings += rule_missing_image(parsed)
    findings += rule_variant_gap(parsed, exclude_rows=family_break_rows)
    # Stable ordering: by row, then by rule registration order, then by SKU.
    rule_order = {r: i for i, r in enumerate(RULES)}
    findings.sort(key=lambda f: (f.row, rule_order.get(f.rule, 999), f.sku or ""))
    return findings
